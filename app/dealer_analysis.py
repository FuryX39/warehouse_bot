"""Сравнение двух Excel-отчётов по заказам дилера (артикул / кол-во / цены)."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from typing import BinaryIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

_HEADER_HINTS = frozenset(
    {
        "артикул",
        "sku",
        "код",
        "наименование",
        "количество",
        "кол-во",
        "кол",
        "qty",
        "quantity",
        "цена",
        "сумма",
        "итого",
    }
)


_QTY_HINTS = ("колич", "кол-во", "кол во", "qty", "quantity")
_UNIT_PRICE_HINTS = ("цена", "price", "за шт", "за единицу")
_TOTAL_HINTS = ("сумм", "итог", "выручк", "оборот", "amount", "total")
_INVOICE_HINTS = ("накладн", "invoice")
_SUMMARY_HINTS = ("итог", "всего", "итого")


@dataclass(frozen=True)
class DealerPeriodLine:
    sku: str
    quantity: float
    unit_price: float
    line_total: float


@dataclass(frozen=True)
class DealerComparisonRow:
    sku: str
    qty_a: float
    qty_b: float
    unit_price_a: float | None
    unit_price_b: float | None
    total_a: float
    total_b: float
    trend: str
    trend_detail: str


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _cell_float(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace("\u00a0", "").replace(" ", "")
    if not s:
        return None
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _row_looks_like_header(cells: tuple[object, object, object, object]) -> bool:
    joined = " ".join(_cell_str(c).lower() for c in cells)
    if not joined:
        return False
    if any(h in joined for h in _HEADER_HINTS):
        return True
    # Первая ячейка текст, вторая не число — вероятно заголовок.
    if cells[0] is not None and _cell_float(cells[1]) is None and _cell_str(cells[0]):
        return bool(re.search(r"[а-яa-z]", _cell_str(cells[0]).lower()))
    return False


def _looks_like_invoice_row(sku_or_name: str) -> bool:
    low = sku_or_name.casefold()
    return any(h in low for h in _INVOICE_HINTS)


def _looks_like_summary_row(sku_or_name: str) -> bool:
    low = sku_or_name.casefold()
    return any(h in low for h in _SUMMARY_HINTS)


def _detect_column_mapping(ws) -> tuple[int, int, int | None, int | None]:
    """
    Возвращает индексы: (sku, qty, unit_price|None, total|None).
    По умолчанию старый формат A-D: артикул, кол-во, цена/шт, сумма.
    Для отчётов дилера с группировкой обычно: A=товар, B=сумма, C=кол-во.
    """
    sku_col = 0
    qty_col = 1
    unit_col: int | None = 2
    total_col: int | None = 3

    for row in ws.iter_rows(min_row=1, max_row=25, max_col=8, values_only=True):
        labels = [(_cell_str(v).casefold()) for v in row]
        if not any(labels):
            continue
        if not any(any(h in c for h in _HEADER_HINTS) for c in labels):
            continue

        qty_idx = next((i for i, c in enumerate(labels) if any(h in c for h in _QTY_HINTS)), None)
        total_idx = next((i for i, c in enumerate(labels) if any(h in c for h in _TOTAL_HINTS)), None)
        unit_idx = next((i for i, c in enumerate(labels) if any(h in c for h in _UNIT_PRICE_HINTS)), None)
        name_idx = next(
            (
                i
                for i, c in enumerate(labels)
                if any(h in c for h in ("артикул", "sku", "код", "товар", "наимен"))
            ),
            None,
        )

        if qty_idx is None:
            continue
        sku_col = name_idx if name_idx is not None else 0
        qty_col = qty_idx
        unit_col = unit_idx
        total_col = total_idx
        return sku_col, qty_col, unit_col, total_col

    return sku_col, qty_col, unit_col, total_col


def _expand_all_grouped_rows(ws) -> None:
    """Снимает hidden/collapsed у строк/колонок с группировкой перед парсингом."""
    for dim in ws.row_dimensions.values():
        try:
            dim.hidden = False
            dim.collapsed = False
        except Exception:
            continue
    for dim in ws.column_dimensions.values():
        try:
            dim.hidden = False
            dim.collapsed = False
        except Exception:
            continue


def _detect_product_outline_levels(ws, sku_col: int, qty_col: int) -> set[int]:
    """
    Определяет уровни строк с товарами на основании структуры групп:
    если есть строки "накладная" на уровне N, то товар обычно на уровне N-1.
    Может вернуть несколько уровней одновременно.
    """
    levels: set[int] = set()
    max_col = max(sku_col, qty_col) + 1
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_col=max_col, values_only=True), start=1):
        sku = _cell_str(row[sku_col] if len(row) > sku_col else None)
        if not sku or not _looks_like_invoice_row(sku):
            continue
        outline = int(getattr(ws.row_dimensions.get(idx), "outlineLevel", 0) or 0)
        if outline > 0:
            levels.add(outline - 1)
    return levels


def parse_dealer_orders_excel(source: bytes | BinaryIO) -> dict[str, DealerPeriodLine]:
    """Читает xlsx: колонки A–D — артикул, кол-во, цена/шт, сумма. Дубликаты SKU суммируются."""
    buf = io.BytesIO(source) if isinstance(source, bytes) else source
    wb = load_workbook(filename=buf, read_only=False, data_only=True)
    try:
        ws = wb.active
        _expand_all_grouped_rows(ws)
        sku_col, qty_col, unit_col, total_col = _detect_column_mapping(ws)
        product_outline_levels = _detect_product_outline_levels(ws, sku_col, qty_col)

        sku_display: dict[str, str] = {}
        aggregated: dict[str, tuple[float, float]] = {}
        max_col = max(sku_col, qty_col, unit_col or 0, total_col or 0) + 1
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_col=max_col, values_only=True), start=1):
            cells = tuple(row) + (None,) * max(0, 4 - len(row))
            if _row_looks_like_header(cells):
                continue
            if product_outline_levels:
                outline = int(getattr(ws.row_dimensions.get(idx), "outlineLevel", 0) or 0)
                if outline not in product_outline_levels:
                    continue

            sku = _cell_str(row[sku_col] if len(row) > sku_col else None)
            if not sku:
                continue
            if _looks_like_invoice_row(sku) or _looks_like_summary_row(sku):
                continue

            qty = _cell_float(row[qty_col] if len(row) > qty_col else None)
            if qty is None or qty <= 0:
                # Нулевые продажи не учитываем.
                continue
            unit_price = _cell_float(row[unit_col] if (unit_col is not None and len(row) > unit_col) else None)
            line_total = _cell_float(row[total_col] if (total_col is not None and len(row) > total_col) else None)
            if line_total is None and unit_price is not None:
                line_total = unit_price * qty
            elif line_total is None:
                line_total = 0.0
            if unit_price is None and qty > 0:
                unit_price = line_total / qty
            key = sku.casefold()
            sku_display[key] = sku
            prev_q, prev_t = aggregated.get(key, (0.0, 0.0))
            aggregated[key] = (prev_q + qty, prev_t + float(line_total))
        out: dict[str, DealerPeriodLine] = {}
        for key, (q, t) in aggregated.items():
            display = sku_display[key]
            out[display] = DealerPeriodLine(
                sku=display,
                quantity=q,
                unit_price=(t / q) if q else 0.0,
                line_total=t,
            )
        return out
    finally:
        wb.close()


def _pct_change(old: float, new: float) -> float | None:
    if old == 0:
        return None if new == 0 else 100.0
    return ((new - old) / old) * 100.0


def _classify_trend(qty_a: float, qty_b: float, total_a: float, total_b: float) -> tuple[str, str]:
    if qty_a <= 0 and qty_b > 0:
        return "Новая позиция", "Не продавалась в первом периоде"
    if qty_b <= 0 and qty_a > 0:
        return "Снята с продаж", "Была в первом периоде, во втором нет"
    if qty_a > 0 and qty_b > 0:
        if qty_b > qty_a:
            rev = ""
            if total_b > total_a:
                rev = ", выручка выросла"
            elif total_b < total_a:
                rev = ", выручка упала при росте кол-ва"
            return "Рост", f"Количество: {qty_a:g} → {qty_b:g}{rev}"
        if qty_b < qty_a:
            rev = ""
            if total_b < total_a:
                rev = ", выручка упала"
            elif total_b > total_a:
                rev = ", выручка выросла при падении кол-ва"
            return "Падение", f"Количество: {qty_a:g} → {qty_b:g}{rev}"
        if total_b > total_a:
            return "Без изменений кол-ва", "Выручка выросла при том же количестве"
        if total_b < total_a:
            return "Без изменений кол-ва", "Выручка упала при том же количестве"
        return "Без изменений", "Количество и выручка совпали"
    return "—", ""


def compare_dealer_periods(
    period_a: dict[str, DealerPeriodLine],
    period_b: dict[str, DealerPeriodLine],
) -> list[DealerComparisonRow]:
    keys_a = {k.casefold(): k for k in period_a}
    keys_b = {k.casefold(): k for k in period_b}
    all_keys = sorted(set(keys_a) | set(keys_b))
    rows: list[DealerComparisonRow] = []
    for key in all_keys:
        sku = keys_b.get(key) or keys_a[key]
        la = period_a.get(keys_a[key]) if key in keys_a else None
        lb = period_b.get(keys_b[key]) if key in keys_b else None
        qty_a = la.quantity if la else 0.0
        qty_b = lb.quantity if lb else 0.0
        total_a = la.line_total if la else 0.0
        total_b = lb.line_total if lb else 0.0
        unit_a = la.unit_price if la and la.quantity else None
        unit_b = lb.unit_price if lb and lb.quantity else None
        trend, detail = _classify_trend(qty_a, qty_b, total_a, total_b)
        rows.append(
            DealerComparisonRow(
                sku=sku,
                qty_a=qty_a,
                qty_b=qty_b,
                unit_price_a=unit_a,
                unit_price_b=unit_b,
                total_a=total_a,
                total_b=total_b,
                trend=trend,
                trend_detail=detail,
            )
        )
    return rows


_TREND_SORT = {
    "Нет продаж": 0,
    "Новая позиция": 1,
    "Падение": 2,
    "Рост": 3,
    "Без изменений кол-ва": 4,
    "Без изменений": 5,
    "—": 6,
}


def build_comparison_stats(rows: list[DealerComparisonRow]) -> dict:
    counts: dict[str, int] = {}
    for r in rows:
        counts[r.trend] = counts.get(r.trend, 0) + 1
    return {
        "sku_count": len(rows),
        "only_period_a": counts.get("Снята с продаж", 0),
        "only_period_b": counts.get("Новая позиция", 0),
        "growth": counts.get("Рост", 0),
        "decline": counts.get("Падение", 0),
        "unchanged": counts.get("Без изменений", 0) + counts.get("Без изменений кол-ва", 0),
        "by_trend": counts,
    }


def export_comparison_excel(
    rows: list[DealerComparisonRow],
    *,
    period_a_label: str,
    period_b_label: str,
) -> bytes:
    label_a = period_a_label or "Период A"
    label_b = period_b_label or "Период B"
    sorted_rows = sorted(
        rows,
        key=lambda r: (_TREND_SORT.get(r.trend, 99), r.sku.casefold()),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Анализ"
    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    band_fill = PatternFill(fill_type="solid", fgColor="F7F9FC")
    total_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    headers = [
        "Артикул",
        "Тренд",
        "Комментарий",
        f"Кол-во ({label_a})",
        f"Кол-во ({label_b})",
        "Δ кол-во",
        "Δ кол-во %",
        f"Цена/шт ({label_a})",
        f"Цена/шт ({label_b})",
        f"Сумма ({label_a})",
        f"Сумма ({label_b})",
        "Δ сумма",
        "Δ сумма %",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    # Закрепляем верхнюю строку (шапку) при прокрутке.
    ws.freeze_panes = "A2"

    for r in sorted_rows:
        dq = r.qty_b - r.qty_a
        dt = r.total_b - r.total_a
        pct_q = _pct_change(r.qty_a, r.qty_b)
        pct_t = _pct_change(r.total_a, r.total_b)
        ws.append(
            [
                r.sku,
                r.trend,
                r.trend_detail,
                r.qty_a or "",
                r.qty_b or "",
                dq if (r.qty_a or r.qty_b) else "",
                round(pct_q, 1) if pct_q is not None else "",
                round(r.unit_price_a, 2) if r.unit_price_a is not None else "",
                round(r.unit_price_b, 2) if r.unit_price_b is not None else "",
                round(r.total_a, 2) if r.total_a else "",
                round(r.total_b, 2) if r.total_b else "",
                round(dt, 2) if (r.total_a or r.total_b) else "",
                round(pct_t, 1) if pct_t is not None else "",
            ]
        )
        row_idx = ws.max_row
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = band_fill

    total_qty_a = sum(r.qty_a for r in sorted_rows)
    total_qty_b = sum(r.qty_b for r in sorted_rows)
    total_rev_a = sum(r.total_a for r in sorted_rows)
    total_rev_b = sum(r.total_b for r in sorted_rows)
    total_qty_delta = total_qty_b - total_qty_a
    total_rev_delta = total_rev_b - total_rev_a
    total_qty_pct = _pct_change(total_qty_a, total_qty_b)
    total_rev_pct = _pct_change(total_rev_a, total_rev_b)
    sold_positions_a = sum(1 for r in sorted_rows if r.qty_a > 0)
    sold_positions_b = sum(1 for r in sorted_rows if r.qty_b > 0)

    ws.append([""] * len(headers))
    ws.append(
        [
            "ИТОГО",
            "",
            (
                f"Свод по всем товарам. Проданных позиций: "
                f"{sold_positions_a} → {sold_positions_b}"
            ),
            round(total_qty_a, 2),
            round(total_qty_b, 2),
            round(total_qty_delta, 2),
            round(total_qty_pct, 1) if total_qty_pct is not None else "",
            "",
            "",
            round(total_rev_a, 2),
            round(total_rev_b, 2),
            round(total_rev_delta, 2),
            round(total_rev_pct, 1) if total_rev_pct is not None else "",
        ]
    )
    total_row_idx = ws.max_row
    for cell in ws[total_row_idx]:
        cell.font = Font(bold=True)
        cell.fill = total_fill

    # Подбираем ширину колонок, чтобы текст читался без ручного расширения.
    for col_cells in ws.columns:
        values = [str(c.value).strip() for c in col_cells if c.value is not None]
        if not values:
            continue
        max_len = max(len(v) for v in values)
        col_letter = col_cells[0].column_letter
        # Ограничиваем, чтобы слишком длинные комментарии не раздували лист.
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def run_dealer_analysis(
    file_a: bytes,
    file_b: bytes,
    *,
    period_a_label: str,
    period_b_label: str,
) -> tuple[list[DealerComparisonRow], dict, bytes]:
    period_a = parse_dealer_orders_excel(file_a)
    period_b = parse_dealer_orders_excel(file_b)
    rows = compare_dealer_periods(period_a, period_b)
    stats = build_comparison_stats(rows)
    stats["period_a_skus"] = len(period_a)
    stats["period_b_skus"] = len(period_b)
    report_xlsx = export_comparison_excel(rows, period_a_label=period_a_label, period_b_label=period_b_label)
    return rows, stats, report_xlsx
