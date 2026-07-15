"""Расчёт объёма, веса и количества грузомест по Excel со SKU и количеством."""

from __future__ import annotations

import io
from collections import OrderedDict
from decimal import Decimal, InvalidOperation, ROUND_CEILING
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

_TEMPLATE_HEADERS = ("Артикул", "Количество")


def build_cargo_place_template() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Товары"
    sheet.append(list(_TEMPLATE_HEADERS))
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8EEF4")
    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 16

    help_sheet = workbook.create_sheet("Описание")
    help_sheet.append(["Заполните артикул товара из каталога и целое количество больше нуля."])
    help_sheet.append(["Повторяющиеся артикулы будут объединены, количества — сложены."])
    help_sheet.column_dimensions["A"].width = 90

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _find_columns(sheet) -> tuple[int, int, int]:
    for row_index, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=min(25, sheet.max_row), values_only=True),
        start=1,
    ):
        sku_column = None
        quantity_column = None
        for column_index, raw in enumerate(row):
            value = _text(raw).casefold().replace("ё", "е")
            if not value:
                continue
            if value in {"артикул", "sku", "ваш sku", "ваш sku *"} or "артикул" in value:
                sku_column = column_index
            if (
                value in {"количество", "кол-во", "кол во", "qty", "quantity"}
                or value.startswith("колич")
            ):
                quantity_column = column_index
        if sku_column is not None and quantity_column is not None:
            return row_index, sku_column, quantity_column
    raise ValueError("Не найдены колонки «Артикул» и «Количество»")


def _parse_quantity(value: object, *, row_index: int) -> int:
    raw = _text(value).replace(",", ".")
    if not raw:
        raise ValueError(f"Строка {row_index}: не указано количество")
    try:
        number = Decimal(raw)
    except InvalidOperation as exc:
        raise ValueError(f"Строка {row_index}: количество должно быть числом") from exc
    if number != number.to_integral_value() or number <= 0:
        raise ValueError(f"Строка {row_index}: количество должно быть целым и больше нуля")
    return int(number)


def parse_cargo_place_workbook(data: bytes) -> list[dict[str, Any]]:
    if not data:
        raise ValueError("Файл пустой")
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Не удалось прочитать Excel-файл (.xlsx)") from exc
    sheet = workbook.active
    if sheet is None:
        raise ValueError("В файле нет листа с товарами")
    header_row, sku_column, quantity_column = _find_columns(sheet)
    merged: OrderedDict[str, dict[str, Any]] = OrderedDict()
    for row_index, row in enumerate(
        sheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        sku = _text(row[sku_column] if sku_column < len(row) else None)
        quantity_raw = row[quantity_column] if quantity_column < len(row) else None
        if not sku and not _text(quantity_raw):
            continue
        if not sku:
            raise ValueError(f"Строка {row_index}: не указан артикул")
        quantity = _parse_quantity(quantity_raw, row_index=row_index)
        key = sku.casefold()
        if key in merged:
            merged[key]["quantity"] += quantity
        else:
            merged[key] = {"sku": sku, "quantity": quantity, "source_row": row_index}
    if not merged:
        raise ValueError("В файле нет товаров для расчёта")
    return list(merged.values())


def _positive_decimal(value: object) -> Decimal | None:
    text = str(value or "").strip().replace(",", ".")
    if not text:
        return None
    try:
        number = Decimal(text)
    except InvalidOperation:
        return None
    return number if number > 0 else None


def _number(value: Decimal | None) -> float | None:
    if value is None:
        return None
    return float(value)


def calculate_cargo_places(
    data: bytes,
    *,
    catalog_repo: Any,
    cargo_place_type_id: int,
) -> dict[str, Any]:
    source_rows = parse_cargo_place_workbook(data)
    cargo_types = catalog_repo.list_cargo_place_types()
    cargo_type = next(
        (item for item in cargo_types if int(item["id"]) == int(cargo_place_type_id)),
        None,
    )
    if cargo_type is None:
        raise ValueError("Выбранный вид грузоместа не найден")
    cargo_volume = _positive_decimal(cargo_type.get("volume_liters"))
    if cargo_volume is None:
        raise ValueError("У выбранного грузоместа некорректные габариты")

    products = catalog_repo.lookup_products_with_metrics_by_skus(
        [row["sku"] for row in source_rows]
    )
    result_rows: list[dict[str, Any]] = []
    total_quantity = 0
    total_volume = Decimal("0")
    total_weight = Decimal("0")
    volume_complete = True
    weight_complete = True

    for source in source_rows:
        sku = source["sku"]
        quantity = int(source["quantity"])
        total_quantity += quantity
        product = products.get(sku.casefold())
        if product is None:
            volume_complete = False
            weight_complete = False
            result_rows.append(
                {
                    "product_id": None,
                    "sku": sku,
                    "name": "",
                    "quantity": quantity,
                    "length_mm": "",
                    "width_mm": "",
                    "height_mm": "",
                    "unit_volume_liters": None,
                    "total_volume_liters": None,
                    "unit_weight_kg": None,
                    "total_weight_kg": None,
                    "missing_fields": ["product"],
                    "error": "Товар с таким артикулом не найден в каталоге",
                }
            )
            continue

        length = _positive_decimal(product.get("length_mm"))
        width = _positive_decimal(product.get("width_mm"))
        height = _positive_decimal(product.get("height_mm"))
        weight = _positive_decimal(product.get("weight_kg"))
        missing_fields: list[str] = []
        for key, value in (("length_mm", length), ("width_mm", width), ("height_mm", height)):
            if value is None:
                missing_fields.append(key)
        if weight is None:
            missing_fields.append("weight_kg")

        unit_volume = None
        row_volume = None
        if length is not None and width is not None and height is not None:
            unit_volume = length * width * height / Decimal("1000000")
            row_volume = unit_volume * quantity
            total_volume += row_volume
        else:
            volume_complete = False

        row_weight = None
        if weight is not None:
            row_weight = weight * quantity
            total_weight += row_weight
        else:
            weight_complete = False

        result_rows.append(
            {
                "product_id": int(product["id"]),
                "sku": product.get("sku") or sku,
                "name": product.get("name") or "",
                "quantity": quantity,
                "length_mm": product.get("length_mm") or "",
                "width_mm": product.get("width_mm") or "",
                "height_mm": product.get("height_mm") or "",
                "unit_volume_liters": _number(unit_volume),
                "total_volume_liters": _number(row_volume),
                "unit_weight_kg": _number(weight),
                "total_weight_kg": _number(row_weight),
                "missing_fields": missing_fields,
                "error": (
                    "Заполните недостающие ОВХ и вес"
                    if missing_fields
                    else ""
                ),
            }
        )

    cargo_place_count = None
    if volume_complete:
        cargo_place_count = int(
            (total_volume / cargo_volume).to_integral_value(rounding=ROUND_CEILING)
        )

    return {
        "cargo_place_type": cargo_type,
        "rows": result_rows,
        "total_sku_rows": len(result_rows),
        "total_quantity": total_quantity,
        "total_volume_liters": _number(total_volume),
        "total_weight_kg": _number(total_weight),
        "volume_complete": volume_complete,
        "weight_complete": weight_complete,
        "cargo_place_count": cargo_place_count,
    }
