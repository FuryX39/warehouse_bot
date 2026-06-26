"""Массовое добавление и изменение штрихкодов каталога из Excel."""

from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.catalog_repository import CatalogRepository

_TEMPLATE_HEADERS = (
    "Артикул",
    "Код",
    "Название товара",
    "Штрихкод*",
    "Название ШК",
    "Группа ШК",
)

_EXAMPLE_ROW = (
    "ART-001",
    "00001",
    "Пример товара",
    "4601234567890",
    "ШК Озон",
    "Озон",
)

_ERROR_HEADER = "Ошибка"
_ROW_HEADER = "Строка в файле"


@dataclass(frozen=True)
class BarcodeImportResult:
    created: int
    updated: int
    failed: int
    error_report: bytes | None
    total_rows: int


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value).strip()
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def build_barcode_import_template(catalog_repo: CatalogRepository) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Штрихкоды"
    ws.append(list(_TEMPLATE_HEADERS))

    products = catalog_repo.list_products({})
    has_rows = False
    for product in products:
        full = catalog_repo.get_product(int(product.id))
        if full is None:
            continue
        if full.barcodes:
            for item in full.barcodes:
                ws.append(
                    [
                        full.sku or "",
                        full.code or "",
                        full.name or "",
                        item.get("barcode") or "",
                        item.get("label") or "",
                        item.get("group") or "",
                    ]
                )
                has_rows = True
        else:
            ws.append([full.sku or "", full.code or "", full.name or "", "", "", ""])
            has_rows = True

    if not has_rows:
        ws.append(list(_EXAMPLE_ROW))

    for col in range(1, len(_TEMPLATE_HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8EEF4")

    ref = wb.create_sheet("Подсказки")
    ref.append(["Поля"])
    ref.append(["Укажите артикул или код товара и штрихкод в каждой строке."])
    ref.append(["Если штрихкод уже есть у товара — обновятся название и/или группа."])
    ref.append(["Если штрихкода у товара ещё нет — он будет добавлен."])
    ref.append(["Пустые «Название ШК» и «Группа ШК» при обновлении не меняют соответствующие поля."])
    ref.append(["Колонка «Название товара» только для удобства, при загрузке не используется."])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _row_is_empty(values: list[str]) -> bool:
    return not any((values[0].strip(), values[1].strip(), values[3].strip()))


def _parse_data_rows(sheet) -> list[tuple[int, list[str]]]:
    rows: list[tuple[int, list[str]]] = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = [_cell_str(cell) for cell in row[: len(_TEMPLATE_HEADERS)]]
        while len(values) < len(_TEMPLATE_HEADERS):
            values.append("")
        if _row_is_empty(values):
            continue
        if row_idx == 2 and values[0].strip().upper() == "ART-001" and values[3].strip() == "4601234567890":
            if not values[4].strip() and not values[5].strip():
                continue
        rows.append((row_idx, values))
    return rows


def _build_error_report(failed_rows: list[tuple[int, list[str], str]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ошибки"
    headers = [_ROW_HEADER, *_TEMPLATE_HEADERS, _ERROR_HEADER]
    ws.append(headers)
    for row_idx, values, error in failed_rows:
        ws.append([row_idx, *values, error])
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
    error_col = len(headers)
    fill = PatternFill("solid", fgColor="FFEBEE")
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=error_col).fill = fill

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _resolve_product(
    by_sku: dict[str, dict[str, Any]],
    by_code: dict[str, dict[str, Any]],
    *,
    sku: str,
    code: str,
) -> dict[str, Any]:
    sku = sku.strip()
    code = code.strip()
    if not sku and not code:
        raise ValueError("Укажите артикул или код товара")

    found: dict[int, dict[str, Any]] = {}
    if sku:
        product = by_sku.get(sku.casefold())
        if product is None:
            raise ValueError(f"Товар с артикулом «{sku}» не найден")
        found[int(product["id"])] = product
    if code:
        product = by_code.get(code.casefold())
        if product is None:
            raise ValueError(f"Товар с кодом «{code}» не найден")
        found[int(product["id"])] = product
    if len(found) > 1:
        raise ValueError("Артикул и код указывают на разные товары")
    return next(iter(found.values()))


def import_barcodes_from_xlsx(catalog_repo: CatalogRepository, data: bytes) -> BarcodeImportResult:
    if not data:
        raise ValueError("Файл пустой")
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Не удалось прочитать Excel-файл (.xlsx)") from exc

    sheet = wb.active
    if sheet is None:
        raise ValueError("В файле нет листа с данными")

    parsed_rows = _parse_data_rows(sheet)
    if not parsed_rows:
        raise ValueError("В файле нет строк со штрихкодами для загрузки")

    by_sku, by_code, _by_barcode = catalog_repo.build_product_import_index()
    created = 0
    updated = 0
    failed_rows: list[tuple[int, list[str], str]] = []

    for row_idx, values in parsed_rows:
        sku, code, _product_name, barcode, label, group = values
        try:
            if not barcode.strip():
                raise ValueError("Укажите штрихкод")
            product = _resolve_product(by_sku, by_code, sku=sku, code=code)
            touch_label = bool(label.strip())
            touch_group = bool(group.strip())
            action = catalog_repo.merge_product_barcode(
                product_id=int(product["id"]),
                barcode=barcode,
                label=label,
                group=group,
                touch_label=touch_label,
                touch_group=touch_group,
            )
            if action == "created":
                created += 1
            else:
                if not touch_label and not touch_group:
                    raise ValueError(
                        "Штрихкод уже есть у товара — укажите название и/или группу для изменения"
                    )
                updated += 1
        except ValueError as exc:
            failed_rows.append((row_idx, values, str(exc)))

    error_report = _build_error_report(failed_rows) if failed_rows else None
    return BarcodeImportResult(
        created=created,
        updated=updated,
        failed=len(failed_rows),
        error_report=error_report,
        total_rows=len(parsed_rows),
    )
