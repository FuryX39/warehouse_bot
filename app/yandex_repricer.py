"""Репрайсер Яндекс Маркет: расчёт цены по карте и рекомендации по виду цен."""

from __future__ import annotations

import io
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import Any, Optional

SELLER_BREAKPOINT = 500
SELLER_TO_SHOWCASE_LOW = 0.8467
SELLER_TO_SHOWCASE_HIGH = 0.6084

# Витрина (руб.) -> коэффициент «цена по карте / витрина».
# Ступени рассчитаны по tables_examples/more_prices.xlsx (207 примеров).
_CARD_TIERS: tuple[tuple[float, float, float], ...] = (
    (0, 280, 0.72),
    (280, 315, 0.74),
    (315, 400, 0.75),
    (400, 500, 0.77),
    (500, 600, 0.79),
    (600, 700, 0.80),
    (700, 800, 0.83),
    (800, 900, 0.86),
    (900, 1050, 0.89),
    (1050, 1150, 0.91),
    (1150, 1225, 0.92),
    (1225, 1250, 0.94),
    (1250, 10_000_000, 0.97),
)

_CARD_HIGHER_THRESHOLD = 1.10  # меняем, если цена по карте выше вида цен на 10%+
_CARD_LOWER_THRESHOLD = 0.99   # меняем, если цена по карте ниже вида цен на 1%+
_MAX_RAISE_CORRECTION = 1.50   # ограничение из-за непредсказуемого соинвеста Яндекса


@dataclass(frozen=True)
class RepricerRowResult:
    row_index: int
    sku: str
    name: str
    seller_price: Optional[float]
    showcase_price: Optional[float]
    estimated_card_price: Optional[int]
    catalog_price: Optional[float]
    recommended_seller_price: Optional[int]
    recommended_old_price: Optional[int]
    updated: bool
    missing_catalog_price: bool
    note: str


@dataclass(frozen=True)
class RepricerResult:
    rows: list[RepricerRowResult]
    stats: dict[str, int]
    workbook_bytes: bytes


def _card_multiplier(showcase: float) -> float:
    for lo, hi, mult in _CARD_TIERS:
        if lo <= showcase < hi:
            return mult
    return 0.97


def showcase_from_seller(seller: float) -> int:
    if seller < SELLER_BREAKPOINT:
        return round(seller * SELLER_TO_SHOWCASE_LOW)
    return round(seller * SELLER_TO_SHOWCASE_HIGH)


def card_from_showcase(showcase: float) -> int:
    return round(showcase * _card_multiplier(showcase))


def card_from_seller(seller: int, *, showcase_ratio: Optional[float] = None) -> int:
    if showcase_ratio is None:
        showcase = showcase_from_seller(seller)
    else:
        showcase = round(seller * showcase_ratio)
    return card_from_showcase(showcase)


def _raise_showcase_elasticity(showcase_ratio: float) -> float:
    """Доля повышения цены продавца, доходящая до витрины.

    При сильном соинвесте Яндекс увеличивает свою скидку вместе с ценой
    продавца, поэтому витрина растёт существенно медленнее. Коэффициенты
    рассчитаны по 196 фактическим изменениям из tables_examples/new/.
    """
    if showcase_ratio < 0.525:
        return 0.35
    if showcase_ratio < 0.60:
        return 0.45
    if showcase_ratio < 0.70:
        return 0.35
    return 1.0


def _project_showcase(
    seller: int,
    *,
    current_seller: Optional[int],
    current_showcase: Optional[float],
    showcase_ratio: Optional[float],
) -> int:
    if (
        current_seller is not None
        and current_seller > 0
        and current_showcase is not None
        and current_showcase > 0
        and seller > current_seller
    ):
        ratio = current_showcase / current_seller
        elasticity = _raise_showcase_elasticity(ratio)
        return round(current_showcase * (seller / current_seller) ** elasticity)
    if showcase_ratio is not None:
        return round(seller * showcase_ratio)
    return showcase_from_seller(seller)


def _seller_search_upper(
    target_card: int,
    showcase_ratio: Optional[float] = None,
    current_seller: Optional[int] = None,
) -> int:
    ratio = showcase_ratio or SELLER_TO_SHOWCASE_HIGH
    return max(
        target_card * 3,
        int(target_card / (ratio * 0.72)) + 500,
        int(current_seller or 0) * 10,
    )


def seller_from_showcase(showcase: float) -> int:
    candidates: list[int] = []
    via_low = round(showcase / SELLER_TO_SHOWCASE_LOW)
    if via_low < SELLER_BREAKPOINT and showcase_from_seller(via_low) == round(showcase):
        candidates.append(via_low)
    via_high = round(showcase / SELLER_TO_SHOWCASE_HIGH)
    if via_high >= SELLER_BREAKPOINT and showcase_from_seller(via_high) == round(showcase):
        candidates.append(via_high)
    if not candidates:
        via_high_f = showcase / SELLER_TO_SHOWCASE_HIGH
        if via_high_f >= SELLER_BREAKPOINT:
            candidates.append(round(via_high_f))
        else:
            candidates.append(round(showcase / SELLER_TO_SHOWCASE_LOW))
    return min(candidates)


def showcase_from_target_card(target_card: float, *, at_least: bool = False) -> int:
    target = max(1, round(target_card))
    upper = max(target + 1, int(target / 0.71) + 500)
    best_showcase = target
    best_diff = 10**9
    for showcase in range(max(1, target - 3), upper):
        card = card_from_showcase(showcase)
        if at_least:
            if card >= target:
                return showcase
            continue
        diff = abs(card - target)
        if diff < best_diff:
            best_diff = diff
            best_showcase = showcase
        if diff == 0:
            break
    return best_showcase


def seller_from_target_card(
    target_card: float,
    *,
    current_seller: Optional[int] = None,
    current_showcase: Optional[float] = None,
    raise_price: bool = True,
    showcase_ratio: Optional[float] = None,
) -> int:
    """Подбор цены продавца с учётом соинвеста и скидки по карте.

    При поднятии (raise_price=True) — минимальная цена продавца, при которой
    цена по карте не ниже цели. При снижении — максимальная, при которой не выше.
    При снижении сохраняется фактический коэффициент «витрина / продавец».
    При повышении учитывается наблюдаемое снижение отдачи соинвеста.
    """
    target = max(1, round(target_card))
    if raise_price:
        start = max(1, int(current_seller or 1))
        upper = _seller_search_upper(target, showcase_ratio, current_seller)
        if showcase_ratio is not None and current_showcase is not None:
            linear_recommended = upper
            for seller in range(start, upper + 1):
                if card_from_seller(seller, showcase_ratio=showcase_ratio) >= target:
                    linear_recommended = seller
                    break
            upper = min(upper, round(linear_recommended * _MAX_RAISE_CORRECTION))
        for seller in range(start, upper + 1):
            showcase = _project_showcase(
                seller,
                current_seller=current_seller,
                current_showcase=current_showcase,
                showcase_ratio=showcase_ratio,
            )
            if card_from_showcase(showcase) >= target:
                return seller
        return upper

    start = (
        int(current_seller)
        if current_seller is not None
        else _seller_search_upper(target, showcase_ratio, current_seller)
    )
    start = max(1, start)
    for seller in range(start, 0, -1):
        showcase = _project_showcase(
            seller,
            current_seller=current_seller,
            current_showcase=current_showcase,
            showcase_ratio=showcase_ratio,
        )
        if card_from_showcase(showcase) <= target:
            return seller
    return 1


def _parse_number(value: object) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("\u00a0", "").replace(" ", "").replace(",", ".")
    if not text:
        return None
    try:
        return float(Decimal(text))
    except (InvalidOperation, ValueError):
        return None


def _find_header_row(rows: list[tuple]) -> int:
    for idx, row in enumerate(rows[:25]):
        for cell in row:
            text = str(cell or "").casefold()
            if "sku" in text and "ваш" in text:
                return idx
            if text in {"ваш sku *", "ваш sku"}:
                return idx
    raise ValueError("Не найдена строка заголовков (колонка «Ваш SKU»)")


def _column_map(header: tuple) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(header):
        text = str(cell or "").strip().casefold()
        if not text:
            continue
        if "sku" in text:
            mapping["sku"] = idx
        elif text == "цена *" or text.startswith("цена *"):
            mapping["seller"] = idx
        elif "зач" in text and "цен" in text:
            mapping["old_price"] = idx
        elif "минимум" in text and "акци" in text:
            mapping["min_promo"] = idx
        elif "витрин" in text:
            mapping["showcase"] = idx
        elif text == "название товара" or ("название" in text and "товар" in text):
            mapping["name"] = idx
    missing = [key for key in ("sku", "seller", "showcase") if key not in mapping]
    if missing:
        raise ValueError(f"В файле нет обязательных колонок: {', '.join(missing)}")
    if "min_promo" not in mapping and "seller" in mapping:
        mapping["min_promo"] = mapping["seller"] + 1
    if "old_price" not in mapping and "seller" in mapping:
        mapping["old_price"] = mapping["seller"] + 1
    if "name" not in mapping:
        mapping["name"] = mapping["sku"] + 1
    return mapping


def _needs_reprice(card_price: float, catalog_price: float) -> bool:
    if catalog_price <= 0:
        return False
    if card_price >= catalog_price * _CARD_HIGHER_THRESHOLD:
        return True
    if card_price <= catalog_price * _CARD_LOWER_THRESHOLD:
        return True
    return False


def _reprice_note(card_price: float, catalog_price: float, *, updated: bool) -> str:
    if updated:
        if card_price < catalog_price:
            return "цена по карте ниже вида цен на 1% и более"
        return "цена по карте выше вида цен на 10% и более"
    if card_price < catalog_price:
        return "цена по карте ниже, но менее чем на 1%"
    if card_price > catalog_price:
        return "цена по карте выше, но менее чем на 10%"
    return "цена по карте совпадает с видом цен"


def _catalog_price_map(
    catalog_repo: Any,
    price_type_id: int,
) -> dict[str, float]:
    products = catalog_repo.list_products_for_price_type(int(price_type_id), {})
    out: dict[str, float] = {}
    for item in products:
        sku = str(item.get("sku") or "").strip()
        if not sku:
            continue
        parsed = _parse_number(item.get("price"))
        if parsed is None:
            continue
        out[sku.casefold()] = parsed
    return out


def process_yandex_prices_workbook(
    data: bytes,
    *,
    catalog_repo: Any,
    price_type_id: int,
    price_type_name: str = "",
) -> RepricerResult:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Не установлен openpyxl") from exc

    if not data:
        raise ValueError("Файл пустой")

    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    if ws is None:
        raise ValueError("В файле нет листа с данными")

    rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
    header_idx = _find_header_row(rows)
    cols = _column_map(rows[header_idx])
    prices_by_sku = _catalog_price_map(catalog_repo, price_type_id)
    missing_price_label = (
        f"У товара нет вида цен «{price_type_name.strip()}»"
        if price_type_name and price_type_name.strip()
        else "У товара нет выбранного вида цен"
    )

    results: list[RepricerRowResult] = []
    updated_rows_data: list[tuple] = []
    stats = {
        "total_rows": 0,
        "with_showcase": 0,
        "with_catalog_price": 0,
        "updated": 0,
        "skipped_no_showcase": 0,
        "skipped_no_catalog_price": 0,
        "skipped_ok": 0,
        "skipped_same_price": 0,
    }

    max_col_idx = max(cols.values())
    for row_offset, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        sku_raw = row[cols["sku"]] if cols["sku"] < len(row) else None
        sku = str(sku_raw or "").strip()
        if not sku:
            continue
        stats["total_rows"] += 1

        name_cell = row[cols["name"]] if cols["name"] < len(row) else ""
        name = str(name_cell or "").strip()
        seller = _parse_number(row[cols["seller"]] if cols["seller"] < len(row) else None)
        showcase = _parse_number(row[cols["showcase"]] if cols["showcase"] < len(row) else None)
        catalog_price = prices_by_sku.get(sku.casefold())

        estimated_card: Optional[int] = None
        recommended: Optional[int] = None
        recommended_old: Optional[int] = None
        updated = False
        missing_catalog_price = False
        note = ""

        if showcase is None or showcase <= 0:
            stats["skipped_no_showcase"] += 1
            note = "нет цены на витрине"
        else:
            stats["with_showcase"] += 1
            estimated_card = card_from_showcase(showcase)
            if catalog_price is None:
                stats["skipped_no_catalog_price"] += 1
                missing_catalog_price = True
                note = missing_price_label
            else:
                stats["with_catalog_price"] += 1
                if _needs_reprice(float(estimated_card), catalog_price):
                    raise_price = float(estimated_card) < catalog_price
                    current_seller = int(round(seller)) if seller and seller > 0 else None
                    showcase_ratio = (
                        showcase / seller
                        if seller is not None
                        and seller > 0
                        and 0.2 <= showcase / seller <= 1.2
                        else None
                    )
                    recommended = seller_from_target_card(
                        catalog_price,
                        current_seller=current_seller,
                        current_showcase=showcase,
                        raise_price=raise_price,
                        showcase_ratio=showcase_ratio,
                    )
                    if current_seller is not None and recommended == current_seller:
                        stats["skipped_same_price"] += 1
                        stats["skipped_ok"] += 1
                        note = "рекомендуемая цена совпадает с текущей"
                    else:
                        recommended_old = int(recommended) * 2
                        row_values = list(row)
                        if len(row_values) <= max_col_idx:
                            row_values.extend([None] * (max_col_idx + 1 - len(row_values)))
                        row_values[cols["seller"]] = int(recommended)
                        row_values[cols["old_price"]] = recommended_old
                        row_values[cols["min_promo"]] = int(recommended)
                        updated_rows_data.append(tuple(row_values))
                        updated = True
                        stats["updated"] += 1
                        note = _reprice_note(float(estimated_card), catalog_price, updated=True)
                else:
                    stats["skipped_ok"] += 1
                    note = _reprice_note(float(estimated_card), catalog_price, updated=False)

        results.append(
            RepricerRowResult(
                row_index=row_offset,
                sku=sku,
                name=name,
                seller_price=seller,
                showcase_price=showcase,
                estimated_card_price=estimated_card,
                catalog_price=catalog_price,
                recommended_seller_price=recommended,
                recommended_old_price=recommended_old,
                updated=updated,
                missing_catalog_price=missing_catalog_price,
                note=note,
            )
        )

    first_data_row = header_idx + 2
    existing_data_rows = max(0, ws.max_row - first_data_row + 1)
    if existing_data_rows:
        ws.delete_rows(first_data_row, existing_data_rows)
    for offset, row_data in enumerate(updated_rows_data):
        row_num = first_data_row + offset
        for col_idx, value in enumerate(row_data):
            ws.cell(row=row_num, column=col_idx + 1, value=value)

    out = io.BytesIO()
    wb.save(out)
    return RepricerResult(rows=results, stats=stats, workbook_bytes=out.getvalue())
