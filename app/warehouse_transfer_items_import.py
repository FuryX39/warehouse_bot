"""Массовое добавление строк перемещения из Excel."""

from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.catalog_repository import CatalogRepository
from app.warehouse_transfers_repository import _parse_price_optional, _price_from_sum, _sum_from_price

_TEMPLATE_HEADERS = (
    "Артикул",
    "Код",
    "Штрихкод",
    "Количество*",
    "Цена",
    "Сумма",
)

_EXAMPLE_ROW = (
    "ART-001",
    "00001",
    "",
    "2",
    "100.50",
    "",
)

_ERROR_HEADER = "Ошибка"
_ROW_HEADER = "Строка в файле"


@dataclass(frozen=True)
class TransferItemsImportResult:
    items: list[dict[str, Any]]
    added: int
    failed: int
    error_report: bytes | None
    total_rows: int


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value).strip()
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def build_transfer_items_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Товары"
    ws.append(list(_TEMPLATE_HEADERS))
    ws.append(list(_EXAMPLE_ROW))
    for col in range(1, len(_TEMPLATE_HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8EEF4")

    ref = wb.create_sheet("Подсказки")
    ref.append(["Поля"])
    ref.append(["Укажите хотя бы один идентификатор: артикул, код или штрихкод."])
    ref.append(["Количество обязательно, целое число от 1."])
    ref.append(["Цена и сумма необязательны; можно указать одно из них."])
    ref.append(["Строка с примером (2-я) пропускается, если артикул ART-001."])
    ref.append(["При повторе товара в файле количества суммируются."])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _row_is_empty(values: list[str]) -> bool:
    return not any(v.strip() for v in values[:4])


def _parse_data_rows(sheet) -> list[tuple[int, list[str]]]:
    rows: list[tuple[int, list[str]]] = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = [_cell_str(cell) for cell in row[: len(_TEMPLATE_HEADERS)]]
        while len(values) < len(_TEMPLATE_HEADERS):
            values.append("")
        if _row_is_empty(values):
            continue
        if row_idx == 2 and values[0].strip().upper() == "ART-001":
            continue
        rows.append((row_idx, values))
    return rows


def _build_error_report(failed_rows: list[tuple[int, list[str], str]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ошибки"
    headers = [_ROW_HEADER, *_TEMPLATE_HEADERS, _ERROR_HEADER]
    ws.append(headers)
    for row_idx, values, error in failed_rows:
        ws.append([row_idx, *values, error])
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
    error_col = len(headers)
    fill = PatternFill("solid", fgColor="FFEBEE")
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=error_col).fill = fill

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_quantity(raw: str) -> int:
    text = raw.strip().replace(",", ".")
    if not text:
        raise ValueError("Не указано количество")
    try:
        if "." in text:
            num = float(text)
            if num != int(num):
                raise ValueError("Количество должно быть целым числом")
            qty = int(num)
        else:
            qty = int(text)
    except (TypeError, ValueError) as exc:
        raise ValueError("Некорректное количество") from exc
    if qty < 1:
        raise ValueError("Количество должно быть не меньше 1")
    return qty


def _resolve_product(
    by_sku: dict[str, dict[str, Any]],
    by_code: dict[str, dict[str, Any]],
    by_barcode: dict[str, dict[str, Any]],
    *,
    sku: str,
    code: str,
    barcode: str,
) -> dict[str, Any]:
    sku = sku.strip()
    code = code.strip()
    barcode = barcode.strip()
    if not sku and not code and not barcode:
        raise ValueError("Укажите артикул, код или штрихкод")

    found: dict[int, dict[str, Any]] = {}
    if sku:
        product = by_sku.get(sku.casefold())
        if product is None:
            raise ValueError(f"Товар с артикулом «{sku}» не найден")
        found[int(product["id"])] = product
    if code:
        product = by_code.get(code.casefold())
        if product is None:
            raise ValueError(f"Товар с кодом «{code}» не найден")
        found[int(product["id"])] = product
    if barcode:
        product = by_barcode.get(barcode.casefold())
        if product is None:
            raise ValueError(f"Товар со штрихкодом «{barcode}» не найден")
        found[int(product["id"])] = product

    if len(found) > 1:
        raise ValueError("Артикул, код и штрихкод указывают на разные товары")
    return next(iter(found.values()))


def _line_prices(qty: int, price_raw: str, sum_raw: str) -> tuple[str, str]:
    unit_price = _parse_price_optional(price_raw)
    line_sum = _parse_price_optional(sum_raw)
    if unit_price is None and line_sum is not None:
        unit_price = _price_from_sum(line_sum, qty)
    if line_sum is None and unit_price is not None:
        line_sum = _sum_from_price(unit_price, qty)
    return unit_price or "", line_sum or ""


def _merge_items(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[int, dict[str, Any]] = {}
    for item in items:
        pid = int(item["product_id"])
        if pid not in merged:
            merged[pid] = dict(item)
            continue
        row = merged[pid]
        row["quantity"] = int(row["quantity"]) + int(item["quantity"])
        if item.get("unit_price"):
            row["unit_price"] = item["unit_price"]
            row["line_sum"] = _sum_from_price(str(item["unit_price"]), int(row["quantity"]))
        elif item.get("line_sum"):
            row["line_sum"] = item["line_sum"]
            row["unit_price"] = _price_from_sum(str(item["line_sum"]), int(row["quantity"]))
    return list(merged.values())


def import_transfer_items_from_xlsx(
    repo: CatalogRepository, data: bytes
) -> TransferItemsImportResult:
    if not data:
        raise ValueError("Файл пустой")
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Не удалось прочитать Excel-файл (.xlsx)") from exc

    sheet = wb.active
    if sheet is None:
        raise ValueError("В файле нет листа с данными")

    parsed_rows = _parse_data_rows(sheet)
    if not parsed_rows:
        raise ValueError("В файле нет строк с товарами для загрузки")

    by_sku, by_code, by_barcode = repo.build_product_import_index()
    success_items: list[dict[str, Any]] = []
    failed_rows: list[tuple[int, list[str], str]] = []

    for row_idx, values in parsed_rows:
        sku, code, barcode, qty_raw, price_raw, sum_raw = values
        try:
            product = _resolve_product(
                by_sku, by_code, by_barcode, sku=sku, code=code, barcode=barcode
            )
            qty = _parse_quantity(qty_raw)
            unit_price, line_sum = _line_prices(qty, price_raw, sum_raw)
            success_items.append(
                {
                    "product_id": int(product["id"]),
                    "name": product["name"],
                    "sku": product["sku"],
                    "code": product["code"],
                    "image_url": product.get("image_url") or "",
                    "is_kit": bool(product.get("is_kit")),
                    "quantity": qty,
                    "unit_price": unit_price,
                    "line_sum": line_sum,
                }
            )
        except ValueError as exc:
            failed_rows.append((row_idx, values, str(exc)))

    merged = _merge_items(success_items)
    error_report = _build_error_report(failed_rows) if failed_rows else None
    return TransferItemsImportResult(
        items=merged,
        added=len(merged),
        failed=len(failed_rows),
        error_report=error_report,
        total_rows=len(parsed_rows),
    )
