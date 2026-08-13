"""Отчёт WB об издержках на приём платежей (Finance API)."""

from __future__ import annotations

import io
import logging
import threading
import time
import uuid
from calendar import monthrange
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation
from typing import Any, Callable

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.adapters.base import is_value_configured

logger = logging.getLogger(__name__)

FINANCE_BASE = "https://finance-api.wildberries.ru"
_DETAIL_LIMIT = 21100
_MAX_PAGES = 40
_DEFAULT_INTERVAL_SEC = 61.0
_HEADERS = (
    "ID строки",
    "№ отчёта",
    "Дата эквайринга",
    "Банк-эквайер",
    "ИНН",
    "КПП",
    "Дата продажи",
    "SRID",
    "Тип документа",
    "nmId",
    "Сумма продажи",
    "Комиссия эквайринга",
    "НДС с комиссии",
    "Номер СФ",
    "Дата СФ",
    "shkId",
    "Валюта",
)

PostFn = Callable[[str, dict[str, Any]], tuple[int, Any, str]]
ProgressFn = Callable[[str], None]


@dataclass
class AcquiringReportMeta:
    report_id: int
    seller_name: str
    date_from: str
    date_to: str
    create_date: str
    currency: str
    acquiring_fee_sum: str
    acquiring_fee_vat_sum: str


@dataclass
class AcquiringMonthResult:
    year: int
    month: int
    reports: list[AcquiringReportMeta]
    rows: list[dict[str, Any]]
    net_fee: Decimal
    net_vat: Decimal
    sale_rows: int
    return_rows: int
    workbook_bytes: bytes
    filename: str


def parse_year_month(raw: str) -> tuple[int, int]:
    text = str(raw or "").strip()
    parts = text.split("-")
    if len(parts) != 2:
        raise ValueError("Укажите месяц в формате ГГГГ-ММ")
    try:
        year = int(parts[0])
        month = int(parts[1])
    except ValueError as exc:
        raise ValueError("Укажите месяц в формате ГГГГ-ММ") from exc
    if year < 2020 or year > 2100 or month < 1 or month > 12:
        raise ValueError("Некорректный месяц")
    return year, month


def month_bounds(year: int, month: int) -> tuple[date, date]:
    last = monthrange(int(year), int(month))[1]
    return date(int(year), int(month), 1), date(int(year), int(month), last)


def month_title(year: int, month: int) -> str:
    start, end = month_bounds(year, month)
    return f"с {start.strftime('%d.%m.%Y')} по {end.strftime('%d.%m.%Y')}"


def _excel_safe_text(value: object) -> str:
    text = str(value or "")
    return "".join(ch for ch in text if ord(ch) >= 32 or ch in "\t\n\r")


def _parse_money(value: object) -> Decimal | None:
    raw = str(value or "").strip().replace("\u00a0", "").replace(" ", "").replace(",", ".")
    if not raw:
        return None
    try:
        return Decimal(raw)
    except InvalidOperation:
        return None


def _money0(value: object) -> Decimal:
    return _parse_money(value) or Decimal("0")


def _is_return(document_type: object) -> bool:
    return "возврат" in str(document_type or "").casefold()


def _short_date(value: object) -> str:
    text = str(value or "").strip()
    return text[:10] if text else ""


def default_finance_post(token: str, url: str, payload: dict[str, Any]) -> tuple[int, Any, str]:
    response = requests.post(
        url,
        headers={"Authorization": token, "Content-Type": "application/json"},
        json=payload,
        timeout=180,
    )
    body = (response.text or "").strip()
    data: Any = None
    if body:
        try:
            data = response.json()
        except ValueError:
            data = None
    detail = ""
    if isinstance(data, dict):
        detail = str(data.get("detail") or data.get("title") or "").strip()
    if not detail and body and response.status_code >= 400:
        detail = body[:500]
    return response.status_code, data, detail


class FinanceRateLimiter:
    def __init__(
        self,
        interval_sec: float = _DEFAULT_INTERVAL_SEC,
        sleep: Callable[[float], None] = time.sleep,
    ) -> None:
        self.interval_sec = max(0.0, float(interval_sec))
        self._sleep = sleep
        self._lock = threading.Lock()
        self._last = 0.0

    def wait(self) -> None:
        with self._lock:
            if self.interval_sec <= 0 or self._last <= 0:
                self._last = time.monotonic()
                return
            delay = self.interval_sec - (time.monotonic() - self._last)
            if delay > 0:
                self._sleep(delay)
            self._last = time.monotonic()

    def backoff(self, seconds: float) -> None:
        if seconds > 0:
            self._sleep(seconds)


def _post_finance(
    token: str,
    path: str,
    payload: dict[str, Any],
    *,
    limiter: FinanceRateLimiter,
    post_fn: PostFn | None,
    on_progress: ProgressFn | None,
    label: str,
) -> Any:
    url = FINANCE_BASE + path
    last_detail = ""
    for _attempt in range(1, 6):
        limiter.wait()
        if on_progress:
            on_progress(label)
        if post_fn is None:
            status, data, detail = default_finance_post(token, url, payload)
        else:
            status, data, detail = post_fn(url, payload)
        last_detail = detail
        if status == 429:
            if on_progress:
                on_progress("WB ограничил частоту запросов, ждём…")
            extra = 70.0 if limiter.interval_sec > 0 else 0.0
            if extra:
                limiter.backoff(extra)
            continue
        if status == 204:
            return None
        if status == 401:
            raise ValueError(
                "WB отклонил токен для Finance API. Нужен персональный или сервисный токен "
                "с категорией «Финансы»."
            )
        if status >= 400:
            raise ValueError(detail or f"Ошибка WB API ({status})")
        return data
    raise ValueError(last_detail or "WB API не ответил после повторов (429)")


def fetch_acquiring_list(
    token: str,
    date_from: date,
    date_to: date,
    *,
    limiter: FinanceRateLimiter,
    post_fn: PostFn | None = None,
    on_progress: ProgressFn | None = None,
) -> list[AcquiringReportMeta]:
    data = _post_finance(
        token,
        "/api/finance/v1/acquiring/list",
        {
            "dateFrom": date_from.isoformat(),
            "dateTo": date_to.isoformat(),
            "limit": 1000,
            "offset": 0,
        },
        limiter=limiter,
        post_fn=post_fn,
        on_progress=on_progress,
        label="Запрашиваем список отчётов WB…",
    )
    if not data:
        return []
    if not isinstance(data, list):
        raise ValueError("Неожиданный ответ списка отчётов WB")
    out: list[AcquiringReportMeta] = []
    for item in data:
        if not isinstance(item, dict):
            continue
        report_id = int(item.get("reportId") or 0)
        if report_id <= 0:
            continue
        out.append(
            AcquiringReportMeta(
                report_id=report_id,
                seller_name=str(item.get("sellerFinanceName") or "").strip(),
                date_from=_short_date(item.get("dateFrom")),
                date_to=_short_date(item.get("dateTo")),
                create_date=_short_date(item.get("createDate")),
                currency=str(item.get("currency") or "RUB"),
                acquiring_fee_sum=str(item.get("acquiringFeeSum") or ""),
                acquiring_fee_vat_sum=str(item.get("acquiringFeeVatSum") or ""),
            )
        )
    return out


def fetch_acquiring_details(
    token: str,
    report_id: int,
    *,
    limiter: FinanceRateLimiter,
    post_fn: PostFn | None = None,
    on_progress: ProgressFn | None = None,
    page_size: int = _DETAIL_LIMIT,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    rrd_id = 0
    limit = max(1, int(page_size))
    for page in range(1, _MAX_PAGES + 1):
        data = _post_finance(
            token,
            f"/api/finance/v1/acquiring/detailed/{int(report_id)}",
            {"limit": limit, "rrdId": rrd_id},
            limiter=limiter,
            post_fn=post_fn,
            on_progress=on_progress,
            label=f"Скачиваем детализацию отчёта {report_id}, страница {page}…",
        )
        if not data:
            break
        if not isinstance(data, list):
            raise ValueError("Неожиданный ответ детализации WB")
        batch = [row for row in data if isinstance(row, dict)]
        if not batch:
            break
        rows.extend(batch)
        rrd_id = int(batch[-1].get("rrdId") or 0)
        if len(batch) < limit:
            break
    return rows


def _tally(rows: list[dict[str, Any]]) -> tuple[Decimal, Decimal, int, int]:
    net_fee = Decimal("0")
    net_vat = Decimal("0")
    sale_rows = 0
    return_rows = 0
    for row in rows:
        fee = _money0(row.get("acquiringFee"))
        vat = _money0(row.get("acquiringFeeVat"))
        if _is_return(row.get("documentType") or row.get("docTypeName")):
            return_rows += 1
            net_fee -= fee
            net_vat -= vat
        else:
            sale_rows += 1
            net_fee += fee
            net_vat += vat
    return net_fee, net_vat, sale_rows, return_rows


def build_acquiring_workbook(
    *,
    year: int,
    month: int,
    reports: list[AcquiringReportMeta],
    rows: list[dict[str, Any]],
) -> bytes:
    net_fee, net_vat, sale_rows, return_rows = _tally(rows)
    period = month_title(year, month)
    seller = next((item.seller_name for item in reports if item.seller_name), "")
    official_fee = next((item.acquiring_fee_sum for item in reports if item.acquiring_fee_sum), "")
    official_vat = next((item.acquiring_fee_vat_sum for item in reports if item.acquiring_fee_vat_sum), "")

    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="E8EEF4")
    thin = Border(
        left=Side(style="thin", color="D8DEE9"),
        right=Side(style="thin", color="D8DEE9"),
        top=Side(style="thin", color="D8DEE9"),
        bottom=Side(style="thin", color="D8DEE9"),
    )
    ws["A1"] = "Отчёт об издержках на приём платежей"
    ws["A1"].font = Font(bold=True, size=14)
    summary: list[tuple[str, object]] = [
        ("Период", period),
        ("Юридическое лицо", seller),
        ("№ отчёта", ", ".join(str(item.report_id) for item in reports)),
        ("Дата формирования", ", ".join(item.create_date for item in reports if item.create_date)),
        ("Валюта", reports[0].currency if reports else "RUB"),
        ("Строк детализации", len(rows)),
        ("Продажи, строк", sale_rows),
        ("Возвраты, строк", return_rows),
        ("Сумма издержек по эквайрингу (нетто)", net_fee),
        ("В том числе НДС (нетто)", net_vat),
        ("Сумма издержек по данным WB", official_fee),
        ("НДС по данным WB", official_vat),
    ]
    note_row = 3 + len(summary) + 1
    for i, (label, value) in enumerate(summary, start=3):
        ws.cell(i, 1, label).font = bold
        cell = ws.cell(i, 2, float(value) if isinstance(value, Decimal) else value)
        if isinstance(value, Decimal):
            cell.number_format = "#,##0.00"
    ws.cell(
        note_row,
        1,
        "В детализации комиссия и НДС возвратов указаны плюсом. "
        "В сводке — нетто: продажи минус возвраты. Иначе сумма колонки будет больше, чем в кабинете WB.",
    )
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row + 1, end_column=2)
    ws.cell(note_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 36

    ws2 = wb.create_sheet("Детализация")
    for col, title in enumerate(_HEADERS, start=1):
        cell = ws2.cell(1, col, title)
        cell.font = bold
        cell.fill = fill
        cell.border = thin
        cell.alignment = Alignment(wrap_text=True)
    for i, row in enumerate(rows, start=2):
        values = [
            row.get("rrdId"),
            row.get("reportId"),
            _short_date(row.get("acqDate")),
            _excel_safe_text(row.get("acquiringBank")),
            _excel_safe_text(row.get("tin")),
            _excel_safe_text(row.get("taxRegistrationReasonCode")),
            _short_date(row.get("saleDate")),
            _excel_safe_text(row.get("srid")),
            _excel_safe_text(row.get("documentType") or row.get("docTypeName")),
            row.get("nmId"),
            _parse_money(row.get("retailAmount")),
            _parse_money(row.get("acquiringFee")),
            _parse_money(row.get("acquiringFeeVat")),
            _excel_safe_text(row.get("invoiceNumber")),
            _short_date(row.get("invoiceDate")),
            row.get("shkId"),
            _excel_safe_text(row.get("currency")),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws2.cell(i, col, float(value) if isinstance(value, Decimal) else value)
            if isinstance(value, Decimal):
                cell.number_format = "#,##0.00"
    widths = [16, 14, 16, 28, 16, 16, 14, 42, 16, 14, 16, 20, 16, 16, 14, 16, 10]
    for i, width in enumerate(widths, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = width
    ws2.auto_filter.ref = f"A1:Q{max(1, len(rows) + 1)}"
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_month_acquiring_report(
    token: str,
    year: int,
    month: int,
    *,
    limiter: FinanceRateLimiter | None = None,
    post_fn: PostFn | None = None,
    on_progress: ProgressFn | None = None,
    page_size: int = _DETAIL_LIMIT,
) -> AcquiringMonthResult:
    if not is_value_configured(token):
        raise ValueError("Не задан WB_API_TOKEN")
    date_from, date_to = month_bounds(year, month)
    limiter = limiter or FinanceRateLimiter()
    reports = fetch_acquiring_list(
        token,
        date_from,
        date_to,
        limiter=limiter,
        post_fn=post_fn,
        on_progress=on_progress,
    )
    if not reports:
        raise ValueError(f"За {month_title(year, month)} отчётов об издержках на приём платежей нет")
    rows: list[dict[str, Any]] = []
    for report in reports:
        rows.extend(
            fetch_acquiring_details(
                token,
                report.report_id,
                limiter=limiter,
                post_fn=post_fn,
                on_progress=on_progress,
                page_size=page_size,
            )
        )
    if on_progress:
        on_progress("Собираем Excel…")
    workbook = build_acquiring_workbook(year=year, month=month, reports=reports, rows=rows)
    net_fee, net_vat, sale_rows, return_rows = _tally(rows)
    filename = f"wb_acquiring_{year:04d}-{month:02d}.xlsx"
    return AcquiringMonthResult(
        year=year,
        month=month,
        reports=reports,
        rows=rows,
        net_fee=net_fee,
        net_vat=net_vat,
        sale_rows=sale_rows,
        return_rows=return_rows,
        workbook_bytes=workbook,
        filename=filename,
    )


@dataclass
class _Job:
    job_id: str
    status: str
    message: str
    year: int
    month: int
    filename: str = ""
    error: str = ""
    content: bytes | None = None
    stats: dict[str, Any] = field(default_factory=dict)


class AcquiringExportJobs:
    """Один отчёт за раз: у WB Finance лимит 1 запрос в минуту."""

    def __init__(self, token: str, *, interval_sec: float = _DEFAULT_INTERVAL_SEC) -> None:
        self.token = token
        self.interval_sec = interval_sec
        self._lock = threading.Lock()
        self._jobs: dict[str, _Job] = {}
        self._running = False

    def configured(self) -> bool:
        return is_value_configured(self.token)

    def start(self, year: int, month: int) -> str:
        if not self.configured():
            raise ValueError("Не задан WB_API_TOKEN")
        with self._lock:
            if self._running:
                raise RuntimeError("Уже формируется другой отчёт WB. Дождитесь окончания.")
            job_id = uuid.uuid4().hex
            job = _Job(
                job_id=job_id,
                status="running",
                message="Запуск выгрузки…",
                year=year,
                month=month,
            )
            self._jobs[job_id] = job
            self._running = True
        thread = threading.Thread(target=self._run, args=(job_id, year, month), daemon=True)
        thread.start()
        return job_id

    def snapshot(self, job_id: str) -> dict[str, Any]:
        job = self._jobs.get(job_id)
        if job is None:
            raise KeyError(job_id)
        return {
            "job_id": job.job_id,
            "status": job.status,
            "message": job.message,
            "error": job.error,
            "filename": job.filename,
            "stats": job.stats,
        }

    def download(self, job_id: str) -> tuple[bytes, str]:
        job = self._jobs.get(job_id)
        if job is None:
            raise KeyError(job_id)
        if job.status != "done" or not job.content:
            raise ValueError("Отчёт ещё не готов")
        return job.content, job.filename or "wb_acquiring.xlsx"

    def _run(self, job_id: str, year: int, month: int) -> None:
        job = self._jobs[job_id]

        def on_progress(message: str) -> None:
            job.message = message

        try:
            result = build_month_acquiring_report(
                self.token,
                year,
                month,
                limiter=FinanceRateLimiter(self.interval_sec),
                on_progress=on_progress,
            )
            job.content = result.workbook_bytes
            job.filename = result.filename
            job.stats = {
                "report_ids": [item.report_id for item in result.reports],
                "rows": len(result.rows),
                "sale_rows": result.sale_rows,
                "return_rows": result.return_rows,
                "net_fee": str(result.net_fee),
                "net_vat": str(result.net_vat),
            }
            job.status = "done"
            job.message = "Готово"
        except Exception as exc:
            logger.exception("WB acquiring export failed")
            job.status = "error"
            job.error = str(exc) or "Не удалось сформировать отчёт"
            job.message = job.error
        finally:
            with self._lock:
                self._running = False
            self._prune()

    def _prune(self) -> None:
        keep = 8
        if len(self._jobs) <= keep:
            return
        done = [job for job in self._jobs.values() if job.status in {"done", "error"}]
        done.sort(key=lambda job: job.job_id)
        for job in done[:-keep]:
            self._jobs.pop(job.job_id, None)
