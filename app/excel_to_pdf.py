"""Преобразование Excel (.xlsx) в PDF для доп. инструментов панели."""

from __future__ import annotations

import io
import re
from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.pdf_fonts import get_pdf_label_fonts

EXCEL_TO_PDF_PROFILES: dict[str, dict[str, str]] = {
    "free": {
        "title": "Свободный",
        "description": "Вся таблица с листа, с переносом на страницы.",
    },
    "vseinstrumenti": {
        "title": "ВсеИнструменты",
        "description": "Строки с 18-й, столбцы A, B, C, D, G.",
    },
}

_VSEINSTRUMENTI_START_ROW = 18
_VSEINSTRUMENTI_COLUMNS = (1, 2, 3, 4, 7)
_INVALID_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
_MAX_EXCEL_BYTES = 20 * 1024 * 1024


def list_excel_to_pdf_profiles() -> list[dict[str, str]]:
    return [
        {"id": key, "title": meta["title"], "description": meta.get("description", "")}
        for key, meta in EXCEL_TO_PDF_PROFILES.items()
    ]


def excel_to_pdf_download_name(original_filename: str, profile: str) -> str:
    stem = Path(str(original_filename or "table.xlsx")).stem.strip() or "table"
    safe = _INVALID_FILENAME_CHARS.sub("_", stem).strip(" .") or "table"
    if not safe.lower().endswith(".pdf"):
        safe = f"{safe}.pdf"
    return safe


def excel_to_pdf(content: bytes, *, profile: str, original_filename: str = "") -> bytes:
    profile_key = str(profile or "").strip().lower()
    if profile_key not in EXCEL_TO_PDF_PROFILES:
        raise ValueError("Неизвестный тип преобразования")
    if not content:
        raise ValueError("Файл пустой")
    if len(content) > _MAX_EXCEL_BYTES:
        raise ValueError("Excel слишком большой (макс. 20 МБ)")
    name = str(original_filename or "").lower()
    if not (name.endswith(".xlsx") or name.endswith(".xlsm")):
        raise ValueError("Нужен файл Excel (.xlsx)")

    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise ValueError("Не удалось прочитать Excel-файл") from exc

    try:
        worksheet = workbook.active
        if worksheet is None:
            raise ValueError("В файле нет листов")
        if profile_key == "vseinstrumenti":
            matrix = _extract_vseinstrumenti_matrix(worksheet)
        else:
            matrix = _extract_free_matrix(worksheet)
    finally:
        workbook.close()

    if not matrix:
        raise ValueError("На листе нет данных для PDF")
    return _matrix_to_pdf(matrix)


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if isinstance(value, float):
        text = text.rstrip("0").rstrip(".") if "." in text else text
    return text


def _is_empty_row(values: list[str]) -> bool:
    return not any(str(v or "").strip() for v in values)


def _trim_used_matrix(matrix: list[list[str]]) -> list[list[str]]:
    if not matrix:
        return []
    max_cols = max(len(row) for row in matrix)
    normalized = [row + [""] * (max_cols - len(row)) for row in matrix]
    while normalized and _is_empty_row(normalized[-1]):
        normalized.pop()
    if not normalized:
        return []
    max_cols = len(normalized[0])
    while max_cols > 0:
        if all(not str(row[max_cols - 1] or "").strip() for row in normalized):
            max_cols -= 1
            for row in normalized:
                if len(row) > max_cols:
                    row.pop()
        else:
            break
    return normalized


def _extract_free_matrix(worksheet: Worksheet) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in worksheet.iter_rows(values_only=True):
        rows.append([_cell_str(value) for value in row])
    return _trim_used_matrix(rows)


def _extract_vseinstrumenti_matrix(worksheet: Worksheet) -> list[list[str]]:
    if worksheet.max_row < _VSEINSTRUMENTI_START_ROW:
        raise ValueError(f"В файле нет строки {_VSEINSTRUMENTI_START_ROW} с заголовком таблицы")
    header = [
        _cell_str(worksheet.cell(_VSEINSTRUMENTI_START_ROW, col).value)
        for col in _VSEINSTRUMENTI_COLUMNS
    ]
    if _is_empty_row(header):
        raise ValueError("Строка 18 не содержит заголовков таблицы")
    matrix = [header]
    for row_idx in range(_VSEINSTRUMENTI_START_ROW + 1, (worksheet.max_row or 0) + 1):
        values = [_cell_str(worksheet.cell(row_idx, col).value) for col in _VSEINSTRUMENTI_COLUMNS]
        if _is_empty_row(values):
            continue
        matrix.append(values)
    if len(matrix) < 2:
        raise ValueError("После строки 18 нет строк с товарами")
    return matrix


def _paragraph_text(value: str) -> str:
    return escape(str(value or "")).replace("\n", "<br/>")


def _column_weights(matrix: list[list[str]]) -> list[float]:
    ncols = max(len(row) for row in matrix)
    weights = [1.0] * ncols
    for row in matrix:
        for idx, cell in enumerate(row):
            if idx >= ncols:
                continue
            length = len(str(cell or ""))
            if idx == 0:
                weights[idx] = max(weights[idx], 8.0)
            else:
                weights[idx] = max(weights[idx], min(length, 120) ** 0.55 + 2.0)
    return weights


def _column_widths(matrix: list[list[str]], available_width: float) -> list[float]:
    weights = _column_weights(matrix)
    total = sum(weights) or 1.0
    return [available_width * weight / total for weight in weights]


def _matrix_to_pdf(matrix: list[list[str]]) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle

    font_regular, font_bold = get_pdf_label_fonts()
    ncols = max(len(row) for row in matrix)
    pagesize = landscape(A4) if ncols > 5 else A4
    page_width, page_height = pagesize
    margin = 12 * mm
    available_width = page_width - 2 * margin

    body_style = ParagraphStyle(
        "ExcelPdfCell",
        fontName=font_regular,
        fontSize=8,
        leading=10,
        alignment=TA_LEFT,
    )
    header_style = ParagraphStyle(
        "ExcelPdfHeader",
        fontName=font_bold,
        fontSize=8.5,
        leading=11,
        alignment=TA_LEFT,
    )

    table_data: list[list[Any]] = []
    for row_idx, row in enumerate(matrix):
        padded = row + [""] * (ncols - len(row))
        style = header_style if row_idx == 0 else body_style
        table_data.append([Paragraph(_paragraph_text(cell), style) for cell in padded])

    col_widths = _column_widths(matrix, available_width)
    if ncols == 5 and len(matrix[0]) == 5:
        col_widths = [
            available_width * 0.06,
            available_width * 0.16,
            available_width * 0.48,
            available_width * 0.14,
            available_width * 0.16,
        ]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2f7")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("FONTNAME", (0, 0), (-1, 0), font_bold),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=pagesize,
        leftMargin=margin,
        rightMargin=margin,
        topMargin=margin,
        bottomMargin=margin,
        title="Excel export",
    )
    doc.build([table])
    pdf = buf.getvalue()
    if not pdf.startswith(b"%PDF"):
        raise ValueError("Не удалось сформировать PDF")
    return pdf
