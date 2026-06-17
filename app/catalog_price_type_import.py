"""Массовая загрузка цен по виду цен из Excel."""

from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.catalog_repository import CatalogRepository, _parse_price
from app.crm_repository import CrmRepository

_PRICE_TYPE_LABEL = "Вид цен"
_TEMPLATE_HEADERS = (
    "Артикул",
    "Код",
    "Штрихкод",
    "Название",
    "Цена",
)

_EXAMPLE_ROW = (
    "ART-001",
    "00001",
    "",
    "Пример товара",
    "100.50",
)

_ERROR_HEADER = "Ошибка"
_ROW_HEADER = "Строка в файле"


@dataclass(frozen=True)
class PriceTypePricesImportResult:
    price_type_id: int
    price_type_name: str
    created_price_type: bool
    updated: int
    failed: int
    error_report: bytes | None
    total_rows: int


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value).strip()
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def build_price_type_prices_template(catalog_repo: CatalogRepository) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Цены"
    ws.append([_PRICE_TYPE_LABEL + "*", "Розничная"])
    ws.append(list(_TEMPLATE_HEADERS))

    products = catalog_repo.list_products({})
    if products:
        for product in products:
            ws.append(
                [
                    product.sku or "",
                    product.code or "",
                    "",
                    product.name or "",
                    "",
                ]
            )
    else:
        ws.append(list(_EXAMPLE_ROW))

    for col in range(1, len(_TEMPLATE_HEADERS) + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8EEF4")
    ws.cell(row=1, column=1).font = Font(bold=True)

    ref = wb.create_sheet("Подсказки")
    ref.append(["Поля"])
    ref.append(["В ячейке B1 укажите название вида цен."])
    ref.append(["Если вид цен с таким названием уже есть — цены обновятся в нём."])
    ref.append(["Если нет — будет создан новый вид цен."])
    ref.append(["Для каждой строки товара укажите хотя бы один идентификатор: артикул, код или штрихкод."])
    ref.append(["Колонка «Название» только для удобства, при загрузке не используется."])
    ref.append(["Пустая цена — строка пропускается (цена не меняется)."])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _read_price_type_name(sheet) -> str:
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not first_row:
        raise ValueError("Укажите название вида цен в ячейке B1")
    cells = list(first_row)
    label = _cell_str(cells[0] if len(cells) > 0 else None)
    name = _cell_str(cells[1] if len(cells) > 1 else None)
    if label.casefold().replace("*", "").strip() != _PRICE_TYPE_LABEL.casefold():
        raise ValueError(f'В ячейке A1 должно быть «{_PRICE_TYPE_LABEL}»')
    if not name:
        raise ValueError("Укажите название вида цен в ячейке B1")
    return name[:128]


def _row_is_empty(values: list[str]) -> bool:
    return not any(
        (values[0].strip(), values[1].strip(), values[2].strip(), values[4].strip())
    )


def _parse_data_rows(sheet) -> list[tuple[int, list[str]]]:
    rows: list[tuple[int, list[str]]] = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
        values = [_cell_str(cell) for cell in row[: len(_TEMPLATE_HEADERS)]]
        while len(values) < len(_TEMPLATE_HEADERS):
            values.append("")
        if _row_is_empty(values):
            continue
        if row_idx == 3 and values[0].strip().upper() == "ART-001" and not values[4].strip():
            continue
        rows.append((row_idx, values))
    return rows


def _build_error_report(failed_rows: list[tuple[int, list[str], str]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ошибки"
    headers = [_ROW_HEADER, *_TEMPLATE_HEADERS, _ERROR_HEADER]
    ws.append(headers)
    for row_idx, values, error in failed_rows:
        ws.append([row_idx, *values, error])
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
    error_col = len(headers)
    fill = PatternFill("solid", fgColor="FFEBEE")
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=error_col).fill = fill

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _resolve_product(
    by_sku: dict[str, dict[str, Any]],
    by_code: dict[str, dict[str, Any]],
    by_barcode: dict[str, dict[str, Any]],
    *,
    sku: str,
    code: str,
    barcode: str,
) -> dict[str, Any]:
    sku = sku.strip()
    code = code.strip()
    barcode = barcode.strip()
    if not sku and not code and not barcode:
        raise ValueError("Укажите артикул, код или штрихкод")

    found: dict[int, dict[str, Any]] = {}
    if sku:
        product = by_sku.get(sku.casefold())
        if product is None:
            raise ValueError(f"Товар с артикулом «{sku}» не найден")
        found[int(product["id"])] = product
    if code:
        product = by_code.get(code.casefold())
        if product is None:
            raise ValueError(f"Товар с кодом «{code}» не найден")
        found[int(product["id"])] = product
    if barcode:
        product = by_barcode.get(barcode.casefold())
        if product is None:
            raise ValueError(f"Товар со штрихкодом «{barcode}» не найден")
        found[int(product["id"])] = product

    if len(found) > 1:
        raise ValueError("Артикул, код и штрихкод указывают на разные товары")
    return next(iter(found.values()))


def import_price_type_prices_from_xlsx(
    catalog_repo: CatalogRepository,
    crm_repo: CrmRepository,
    data: bytes,
) -> PriceTypePricesImportResult:
    if not data:
        raise ValueError("Файл пустой")
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Не удалось прочитать Excel-файл (.xlsx)") from exc

    sheet = wb.active
    if sheet is None:
        raise ValueError("В файле нет листа с данными")

    price_type_name = _read_price_type_name(sheet)
    price_type, created = crm_repo.get_or_create_price_type_by_name(price_type_name)

    parsed_rows = _parse_data_rows(sheet)
    if not parsed_rows:
        raise ValueError("В файле нет строк с ценами для загрузки")

    by_sku, by_code, by_barcode = catalog_repo.build_product_import_index()
    success_items: list[dict[str, Any]] = []
    failed_rows: list[tuple[int, list[str], str]] = []

    for row_idx, values in parsed_rows:
        sku, code, barcode, _name, price_raw = values
        try:
            if not str(price_raw or "").strip():
                continue
            product = _resolve_product(
                by_sku, by_code, by_barcode, sku=sku, code=code, barcode=barcode
            )
            price = _parse_price(price_raw)
            if price is None:
                continue
            success_items.append(
                {
                    "product_id": int(product["id"]),
                    "price": price,
                }
            )
        except ValueError as exc:
            failed_rows.append((row_idx, values, str(exc)))

    updated = 0
    if success_items:
        updated = catalog_repo.save_prices_for_price_type(int(price_type["id"]), success_items)

    error_report = _build_error_report(failed_rows) if failed_rows else None
    return PriceTypePricesImportResult(
        price_type_id=int(price_type["id"]),
        price_type_name=str(price_type["name"]),
        created_price_type=created,
        updated=updated,
        failed=len(failed_rows),
        error_report=error_report,
        total_rows=len(parsed_rows),
    )
