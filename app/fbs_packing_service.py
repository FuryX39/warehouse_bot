"""Создание задания FBS-упаковки Яндекс и сопоставление с каталогом."""

from __future__ import annotations

from dataclasses import dataclass
from typing import Any

from app.catalog_repository import CatalogRepository
from app.config import Settings
from app.fbs_labels_common import merge_label_pdfs
from app.fbs_packing_repository import (
    FbsPackingJobRow,
    FbsPackingRepository,
    MARKETPLACE_OZON,
    MARKETPLACE_WB,
    MARKETPLACE_YANDEX,
)
from app.google_sheet_write import fbs_list_sheet_title
from app.marking.cis import parse_cis
from app.marking.match import build_gtin_index
from app.yandex_fbs_labels import (
    YandexFbsListRow,
    _export_list_to_google_sheet,
    build_order_box_labels,
    collect_yandex_unit_labels,
    load_yandex_fbs_list_rows,
    normalize_yandex_fbs_substatus,
)
from app.wb_fbs_labels import (
    collect_wb_unit_labels,
    load_wb_fbs_list_rows,
    normalize_wb_fbs_substatus,
)
from app.ozon_fbs_labels import collect_ozon_unit_labels, load_ozon_fbs_list_rows
from app.adapters.yandex_market import YandexMarketAdapter
from app.adapters.wildberries import WildberriesAdapter
from app.adapters.ozon import OzonAdapter


def _bool(value: object, default: bool = True) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y", "on"}:
        return True
    if text in {"0", "false", "no", "n", "off"}:
        return False
    return default


def resolve_catalog_products(
    catalog: CatalogRepository, skus: list[str]
) -> tuple[dict[str, dict[str, Any]], list[str]]:
    by_sku, _by_code, _by_barcode = catalog.build_product_import_index()
    found: dict[str, dict[str, Any]] = {}
    missing: list[str] = []
    seen: set[str] = set()
    for sku in skus:
        key = str(sku or "").strip().casefold()
        if not key or key in seen:
            continue
        seen.add(key)
        product = by_sku.get(key)
        if product is None:
            missing.append(str(sku).strip())
            continue
        found[key] = product
    return found, missing


def lookup_scan_product(
    catalog: CatalogRepository, raw: str
) -> tuple[str, int | None]:
    text = str(raw or "").strip()
    if not text:
        raise ValueError("Пустой штрихкод")
    by_sku, by_code, by_barcode = catalog.build_product_import_index()
    key = text.casefold()
    product = by_barcode.get(key) or by_sku.get(key) or by_code.get(key)
    if product is None:
        raise ValueError("Штрихкод не найден в каталоге")
    return str(product.get("sku") or ""), int(product["id"]) if product.get("id") else None


@dataclass(frozen=True)
class PackingScanResolve:
    sku: str
    product_id: int | None
    cis_raw: str = ""
    cis_key: str = ""
    cis_gtin: str = ""

    @property
    def is_cis(self) -> bool:
        return bool(self.cis_key)


def product_has_marking_gtin(catalog: CatalogRepository, product_id: int | None) -> bool:
    """Товар маркируемый, если есть в GTIN-индексе (явный GTIN или валидный EAN как GTIN-14)."""
    if not product_id:
        return False
    index, _conflicts = build_gtin_index(catalog)
    want = int(product_id)
    return any(ref.product_id == want for ref in index.values())


def resolve_packing_scan(catalog: CatalogRepository, raw: str) -> PackingScanResolve:
    """
    КИЗ → товар по GTIN (или EAN как GTIN-14).
    Обычный EAN/SKU никогда не считается КИЗом и не пишет маркировку.
    """
    text = str(raw or "").strip()
    if not text:
        raise ValueError("Пустой штрихкод")
    record = parse_cis(text)
    if record.ok:
        index, conflicts = build_gtin_index(catalog)
        gtin = record.gtin
        if gtin in conflicts:
            skus = ", ".join(conflicts[gtin])
            raise ValueError(f"GTIN {gtin} конфликт в каталоге ({skus})")
        product = index.get(gtin)
        if product is None:
            raise ValueError(f"GTIN {gtin} не найден в каталоге")
        return PackingScanResolve(
            sku=product.sku,
            product_id=product.product_id,
            cis_raw=record.raw or text,
            cis_key=record.cis or text,
            cis_gtin=gtin,
        )
    sku, product_id = lookup_scan_product(catalog, text)
    return PackingScanResolve(sku=sku, product_id=product_id)


def create_yandex_packing_job(
    *,
    adapter: YandexMarketAdapter,
    catalog: CatalogRepository,
    packing_repo: FbsPackingRepository,
    settings: Settings,
    order_substatus: str,
    build_list: bool,
    item_limit: int | None,
    packer_user_ids: list[int],
    created_by_user_id: int | None,
    require_cis: bool = False,
) -> FbsPackingJobRow:
    substatus = normalize_yandex_fbs_substatus(order_substatus)
    list_rows, selected_orders, warnings, _available = load_yandex_fbs_list_rows(
        adapter,
        substatus=substatus,
        default_stocks_sheet_url=settings.default_stocks_sheet_url,
        google_service_account_file=settings.google_service_account_file,
        fbs_assembly_sheet_name=settings.fbs_assembly_sheet_name,
        assembly_sheet_gid=settings.fbs_assembly_sheet_gid,
        max_units=item_limit,
        apply_assembly=bool(build_list),
    )
    if not list_rows:
        raise ValueError("Нет заказов для задания")

    units, label_warnings = collect_yandex_unit_labels(
        adapter,
        selected_orders,
        list_rows,
        substatus=substatus,
        label_format=settings.yandex_label_format,
        label_rotate_degrees=settings.yandex_label_rotate_degrees,
    )
    warnings.extend(label_warnings)
    units_with_pdf = [unit for unit in units if unit.pdf]
    if not units_with_pdf:
        detail = "; ".join(warnings[:5]) if warnings else "нет PDF"
        raise ValueError(f"Не удалось получить ярлыки: {detail}")

    catalog_by_sku, missing = resolve_catalog_products(
        catalog, [unit.sku for unit in units_with_pdf]
    )
    for sku in missing:
        warnings.append(f"Артикул «{sku}» не найден в каталоге — строка всё равно в задании")

    sheet_url = ""
    sheet_title = fbs_list_sheet_title() if build_list else ""
    merged_pdf: bytes | None = None
    if build_list:
        pdfs = [unit.pdf for unit in units_with_pdf if unit.pdf]
        merged_pdf = merge_label_pdfs(pdfs)
        if merged_pdf is None and len(pdfs) == 1:
            merged_pdf = pdfs[0]
        sheet_url_cfg = (settings.fbs_list_sheet_url or "").strip()
        creds_path = (settings.google_service_account_file or "").strip()
        job_rows = [
            YandexFbsListRow(
                seq=index,
                order_id=unit.order_id,
                sku=unit.sku,
                quantity=1,
                status=substatus,
            )
            for index, unit in enumerate(units_with_pdf, start=1)
        ]
        if sheet_url_cfg and creds_path and job_rows:
            try:
                sheet_url = _export_list_to_google_sheet(
                    spreadsheet_url=sheet_url_cfg,
                    credentials_path=creds_path,
                    sheet_title=sheet_title,
                    list_rows=job_rows,
                    orders=selected_orders,
                    template_sheet_name=(settings.fbs_list_template_sheet or "FBSTemplate").strip(),
                ) or ""
            except Exception as exc:  # noqa: BLE001
                warnings.append(f"Google Таблица: {exc}")
        elif sheet_url_cfg and job_rows and not creds_path:
            warnings.append(
                "Google Таблица: задайте GOOGLE_SERVICE_ACCOUNT_FILE "
                "(JSON service account с доступом к FBS_LIST_SHEET_URL)."
            )
        elif not sheet_url_cfg and job_rows:
            warnings.append(
                "Google Таблица: задайте FBS_LIST_SHEET_URL (ссылка на таблицу для FBS-списков)."
            )

    line_payloads = []
    for seq, unit in enumerate(units_with_pdf, start=1):
        product = catalog_by_sku.get(unit.sku.casefold())
        line_payloads.append(
            {
                "seq": seq,
                "sku": unit.sku,
                "product_id": int(product["id"]) if product else None,
                "product_name": str(product["name"]) if product else "",
                "order_id": unit.order_id,
                "box_id": unit.box_id,
                "place_index": unit.place_index,
                "place_total": unit.place_total,
                "scan_keys": unit.scan_keys(),
                "pdf": unit.pdf,
            }
        )
    return packing_repo.create_job(
        marketplace=MARKETPLACE_YANDEX,
        order_substatus=substatus,
        build_list=bool(build_list),
        require_cis=bool(require_cis),
        created_by_user_id=created_by_user_id,
        packer_user_ids=packer_user_ids,
        sheet_url=sheet_url,
        sheet_title=sheet_title,
        warnings=warnings,
        merged_pdf=merged_pdf,
        lines=line_payloads,
    )


def create_wb_packing_job(
    *,
    adapter: WildberriesAdapter,
    catalog: CatalogRepository,
    packing_repo: FbsPackingRepository,
    order_substatus: str,
    item_limit: int | None,
    packer_user_ids: list[int],
    created_by_user_id: int | None,
    require_cis: bool = False,
    supply_id: str = "",
) -> FbsPackingJobRow:
    substatus = normalize_wb_fbs_substatus(order_substatus)
    list_rows, selected_orders, warnings, _available = load_wb_fbs_list_rows(
        adapter,
        substatus=substatus,
        supply_id=supply_id,
        max_units=item_limit,
    )
    if not list_rows:
        raise ValueError("Нет заказов для задания")

    units, label_warnings, effective_supply = collect_wb_unit_labels(
        adapter,
        selected_orders,
        substatus=substatus,
        supply_id=supply_id if substatus == "READY_TO_SHIP" else "",
    )
    warnings.extend(label_warnings)
    units_with_pdf = [unit for unit in units if unit.pdf]
    if not units_with_pdf:
        detail = "; ".join(warnings[:5]) if warnings else "нет PDF"
        raise ValueError(f"Не удалось получить стикеры WB: {detail}")

    catalog_by_sku, missing = resolve_catalog_products(
        catalog, [unit.sku for unit in units_with_pdf]
    )
    for sku in missing:
        warnings.append(f"Артикул «{sku}» не найден в каталоге — строка всё равно в задании")

    pdfs = [unit.pdf for unit in units_with_pdf if unit.pdf]
    merged_pdf = merge_label_pdfs(pdfs)
    if merged_pdf is None and len(pdfs) == 1:
        merged_pdf = pdfs[0]

    line_payloads = []
    for seq, unit in enumerate(units_with_pdf, start=1):
        product = catalog_by_sku.get(unit.sku.casefold())
        line_payloads.append(
            {
                "seq": seq,
                "sku": unit.sku,
                "product_id": int(product["id"]) if product else None,
                "product_name": str(product["name"]) if product else "",
                "order_id": unit.order_id,
                "box_id": None,
                "place_index": 1,
                "place_total": 1,
                "scan_keys": unit.scan_keys(),
                "pdf": unit.pdf,
            }
        )
    return packing_repo.create_job(
        marketplace=MARKETPLACE_WB,
        order_substatus=substatus,
        build_list=False,
        require_cis=bool(require_cis),
        supply_id=effective_supply,
        created_by_user_id=created_by_user_id,
        packer_user_ids=packer_user_ids,
        warnings=warnings,
        merged_pdf=merged_pdf,
        lines=line_payloads,
    )


def create_ozon_packing_job(
    *,
    adapter: OzonAdapter,
    catalog: CatalogRepository,
    packing_repo: FbsPackingRepository,
    settings: Settings,
    packer_user_ids: list[int],
    created_by_user_id: int | None,
    require_cis: bool = False,
    first_posting: str = "",
    last_posting: str = "",
) -> FbsPackingJobRow:
    list_rows, _postings, warnings, _available = load_ozon_fbs_list_rows(
        adapter,
        first_posting=first_posting,
        last_posting=last_posting,
    )
    if not list_rows:
        raise ValueError("Нет отправлений Ozon для задания")

    units, label_warnings = collect_ozon_unit_labels(
        adapter,
        list_rows,
        label_rotate_degrees=settings.ozon_label_rotate_degrees,
    )
    warnings.extend(label_warnings)
    units_with_pdf = [unit for unit in units if unit.pdf]
    if not units_with_pdf:
        detail = "; ".join(warnings[:5]) if warnings else "нет PDF"
        raise ValueError(f"Не удалось получить этикетки Ozon: {detail}")

    catalog_by_sku, missing = resolve_catalog_products(
        catalog, [unit.sku for unit in units_with_pdf]
    )
    for sku in missing:
        warnings.append(f"Артикул «{sku}» не найден в каталоге — строка всё равно в задании")

    pdfs = [unit.pdf for unit in units_with_pdf if unit.pdf]
    merged_pdf = merge_label_pdfs(pdfs)
    if merged_pdf is None and len(pdfs) == 1:
        merged_pdf = pdfs[0]

    line_payloads = []
    for seq, unit in enumerate(units_with_pdf, start=1):
        product = catalog_by_sku.get(unit.sku.casefold())
        line_payloads.append(
            {
                "seq": seq,
                "sku": unit.sku,
                "product_id": int(product["id"]) if product else None,
                "product_name": str(product["name"]) if product else "",
                "order_id": unit.posting_number,
                "box_id": None,
                "place_index": 1,
                "place_total": 1,
                "scan_keys": unit.scan_keys(),
                "pdf": unit.pdf,
            }
        )
    return packing_repo.create_job(
        marketplace=MARKETPLACE_OZON,
        order_substatus="awaiting_deliver",
        build_list=False,
        require_cis=bool(require_cis),
        created_by_user_id=created_by_user_id,
        packer_user_ids=packer_user_ids,
        warnings=warnings,
        merged_pdf=merged_pdf,
        lines=line_payloads,
    )


def build_packing_marking_xlsx(job: FbsPackingJobRow) -> bytes:
    """Excel на лету: полный список строк и только строки с КИЗ."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    from app.marking.cis import replace_gs_for_excel

    wb = Workbook()
    headers = ("Заказ", "SKU", "КИЗ")
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E8EEF4")

    def write_sheet(ws, lines: list) -> None:
        ws.append(list(headers))
        for col in range(1, 4):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
        for line in lines:
            cis = replace_gs_for_excel(line.cis_raw or line.cis_key or "")
            ws.append([line.order_display, line.sku, cis])
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 48

    full = wb.active
    full.title = "Все строки"
    ordered = sorted(job.lines, key=lambda item: (item.seq, item.id))
    write_sheet(full, ordered)

    with_cis = wb.create_sheet("С КИЗ")
    write_sheet(with_cis, [line for line in ordered if line.cis_key])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def list_rows_payload(list_rows: list[YandexFbsListRow], orders) -> list[dict[str, Any]]:
    displays = build_order_box_labels(list_rows, orders)
    return [
        {
            "seq": row.seq,
            "sku": row.sku,
            "quantity": row.quantity,
            "order_id": row.order_id,
            "order_display": display,
            "posting_number": row.order_id,
        }
        for row, display in zip(list_rows, displays)
    ]
