"""Массовый импорт товаров каталога из Excel."""

from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.catalog_repository import CatalogRepository

_TEMPLATE_HEADERS = (
    "Название*",
    "Артикул*",
    "Код*",
    "Внешний код",
    "Описание",
    "Ссылка на изображение",
    "Группа",
    "Страна",
    "Единица измерения",
    "Вес",
    "Ширина, мм",
    "Высота, мм",
    "Длина, мм",
    "Объём, л",
    "Тип маркировки",
    "Штрихкоды",
)

_EXAMPLE_ROW = (
    "Пример товара",
    "ART-001",
    "00001",
    "",
    "Краткое описание",
    "",
    "",
    "Россия",
    "шт",
    "",
    "100",
    "50",
    "20",
    "",
    "Не подлежит маркировке",
    "4601234567890",
)

_ERROR_HEADER = "Ошибка"
_ROW_HEADER = "Строка в файле"


@dataclass(frozen=True)
class CatalogImportResult:
    created: int
    updated: int
    failed: int
    error_report: bytes | None
    total_rows: int


def build_products_export(repo: CatalogRepository, filters: dict[str, str]) -> bytes:
    data = repo.list_products_for_export(filters)
    products = data.get("products") or []
    price_types = data.get("price_types") or []
    extra_headers = ["Тип", "Состав комплекта"]
    price_headers = [f"Цена: {item.get('name') or ''}" for item in price_types]
    headers = [*_TEMPLATE_HEADERS, *extra_headers, *price_headers]

    wb = Workbook()
    ws = wb.active
    ws.title = "Товары"
    ws.append(headers)

    for product in products:
        barcodes = "; ".join(
            str(item.get("barcode") or "").strip()
            for item in (product.get("barcodes") or [])
            if str(item.get("barcode") or "").strip()
        )
        components = "; ".join(
            f"{item.get('sku') or item.get('name') or ''} × {int(item.get('quantity') or 1)}"
            for item in (product.get("components") or [])
        )
        prices = product.get("prices") or {}
        ws.append(
            [
                product.get("name") or "",
                product.get("sku") or "",
                product.get("code") or "",
                product.get("external_code") or "",
                product.get("description") or "",
                product.get("image_url") or "",
                product.get("group_name") or "",
                product.get("country") or "",
                product.get("unit_name") or "",
                product.get("weight") or "",
                product.get("width_mm") or "",
                product.get("height_mm") or "",
                product.get("length_mm") or "",
                product.get("volume") or "",
                product.get("marking_type_name") or "",
                barcodes,
                "Комплект" if product.get("is_kit") else "Товар",
                components,
                *[prices.get(int(item["id"]), "") for item in price_types],
            ]
        )

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8EEF4")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        1: 42,
        2: 20,
        3: 14,
        4: 18,
        5: 45,
        6: 36,
        7: 24,
        8: 16,
        9: 18,
        15: 25,
        16: 34,
        17: 12,
        18: 40,
    }
    for column, width in widths.items():
        ws.column_dimensions[ws.cell(row=1, column=column).column_letter].width = width
    for column in range(19, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=column).column_letter].width = 18

    note = wb.create_sheet("Описание")
    note.append(["Массовая выгрузка каталога"])
    note["A1"].font = Font(bold=True)
    note.append(["Первые 16 колонок совместимы с массовой загрузкой товаров."])
    note.append(["Комплекты выгружаются для просмотра, но массовая загрузка их не изменяет."])
    note.append(["Колонки типа, состава комплекта и цен носят информационный характер."])
    note.column_dimensions["A"].width = 85

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value).strip()
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _split_barcodes(raw: str) -> list[str]:
    text = raw.strip()
    if not text:
        return []
    parts: list[str] = []
    for chunk in text.replace(",", ";").split(";"):
        code = chunk.strip()
        if code:
            parts.append(code)
    return parts


def _default_ref_id(items: list[dict[str, Any]]) -> int | None:
    for item in items:
        if item.get("is_default"):
            return int(item["id"])
    if items:
        return int(items[0]["id"])
    return None


def _resolve_ref_id(
    raw: str,
    items: list[dict[str, Any]],
    *,
    label: str,
    use_default: bool,
) -> int | None:
    name = raw.strip()
    if not name:
        return _default_ref_id(items) if use_default else None
    lowered = name.casefold()
    for item in items:
        if str(item.get("name") or "").strip().casefold() == lowered:
            return int(item["id"])
    raise ValueError(f"{label} «{name}» не найден в справочнике")


def build_import_template(meta: dict[str, list[dict[str, Any]]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Товары"
    ws.append(list(_TEMPLATE_HEADERS))
    ws.append(list(_EXAMPLE_ROW))
    for col in range(1, len(_TEMPLATE_HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8EEF4")

    ref = wb.create_sheet("Справочники")
    ref.append(["Группы"])
    for group in meta.get("groups") or []:
        ref.append([group.get("name") or ""])
    ref.append([])
    ref.append(["Единицы измерения"])
    for unit in meta.get("units") or []:
        suffix = " (по умолчанию)" if unit.get("is_default") else ""
        ref.append([f"{unit.get('name') or ''}{suffix}"])
    ref.append([])
    ref.append(["Типы маркировки"])
    for marking in meta.get("marking_types") or []:
        suffix = " (по умолчанию)" if marking.get("is_default") else ""
        ref.append([f"{marking.get('name') or ''}{suffix}"])
    ref.append([])
    ref.append(["Подсказки"])
    ref.append(["Поля со звёздочкой (*) обязательны."])
    ref.append(["Штрихкоды — через точку с запятой, формат Code128."])
    ref.append(["Объём можно не указывать: посчитается из ширины, высоты и длины (мм) в литрах."])
    ref.append(["Если артикул уже есть в каталоге, строка обновит существующий товар."])
    ref.append(["Пустая колонка «Штрихкоды» при обновлении не меняет штрихкоды товара."])
    ref.append(["Строка с примером (2-я) при загрузке пропускается, если артикул ART-001."])
    ref.append(["Комплекты через этот импорт добавить нельзя."])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _row_is_empty(values: list[str]) -> bool:
    return not any(v.strip() for v in values[:3])


def _parse_data_rows(sheet) -> list[tuple[int, list[str]]]:
    rows: list[tuple[int, list[str]]] = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = [_cell_str(cell) for cell in row[: len(_TEMPLATE_HEADERS)]]
        while len(values) < len(_TEMPLATE_HEADERS):
            values.append("")
        if _row_is_empty(values):
            continue
        if row_idx == 2 and values[1].strip().upper() == "ART-001":
            continue
        rows.append((row_idx, values))
    return rows


def _build_error_report(failed_rows: list[tuple[int, list[str], str]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ошибки"
    headers = [_ROW_HEADER, *_TEMPLATE_HEADERS, _ERROR_HEADER]
    ws.append(headers)
    for row_idx, values, error in failed_rows:
        ws.append([row_idx, *values, error])
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
    error_col = len(headers)
    fill = PatternFill("solid", fgColor="FFEBEE")
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=error_col).fill = fill

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def import_products_from_xlsx(repo: CatalogRepository, data: bytes) -> CatalogImportResult:
    if not data:
        raise ValueError("Файл пустой")
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Не удалось прочитать Excel-файл (.xlsx)") from exc

    sheet = wb.active
    if sheet is None:
        raise ValueError("В файле нет листа с данными")

    meta = repo.get_meta()
    parsed_rows = _parse_data_rows(sheet)
    if not parsed_rows:
        raise ValueError("В файле нет строк с товарами для загрузки")

    by_sku, by_code, _ = repo.build_product_import_index()
    seen_skus: dict[str, int] = {}
    seen_codes: dict[str, int] = {}
    seen_barcodes: dict[str, int] = {}
    created = 0
    updated = 0
    failed_rows: list[tuple[int, list[str], str]] = []

    for row_idx, values in parsed_rows:
        (
            name,
            sku,
            code,
            external_code,
            description,
            image_url,
            group_name,
            country,
            unit_name,
            weight,
            width_mm,
            height_mm,
            length_mm,
            volume,
            marking_name,
            barcodes_raw,
        ) = values

        try:
            if not name:
                raise ValueError("Не указано название")
            if not sku:
                raise ValueError("Не указан артикул")
            if not code:
                raise ValueError("Не указан код")

            sku_key = sku.casefold()
            if sku_key in seen_skus:
                raise ValueError(
                    f"Дублирующийся артикул в файле (уже в строке {seen_skus[sku_key]})"
                )
            code_key = code.casefold()
            if code_key in seen_codes:
                raise ValueError(
                    f"Дублирующийся код в файле (уже в строке {seen_codes[code_key]})"
                )

            barcodes = _split_barcodes(barcodes_raw)
            for barcode in barcodes:
                bc_key = barcode.casefold()
                if bc_key in seen_barcodes:
                    raise ValueError(
                        f"Дублирующийся штрихкод «{barcode}» в файле "
                        f"(уже в строке {seen_barcodes[bc_key]})"
                    )

            group_id = _resolve_ref_id(
                group_name, meta.get("groups") or [], label="Группа", use_default=False
            )
            unit_id = _resolve_ref_id(
                unit_name, meta.get("units") or [], label="Единица измерения", use_default=True
            )
            marking_type_id = _resolve_ref_id(
                marking_name,
                meta.get("marking_types") or [],
                label="Тип маркировки",
                use_default=True,
            )

            payload = {
                "is_kit": False,
                "name": name,
                "sku": sku,
                "code": code,
                "external_code": external_code,
                "description": description,
                "image_url": image_url,
                "group_id": group_id,
                "country": country,
                "unit_id": unit_id,
                "weight": weight,
                "width_mm": width_mm,
                "height_mm": height_mm,
                "length_mm": length_mm,
                "volume": volume,
                "marking_type_id": marking_type_id,
                "components": [],
            }
            if barcodes:
                payload["barcodes"] = barcodes

            existing = by_sku.get(sku_key)
            if existing:
                if existing.get("is_kit"):
                    raise ValueError("Комплект нельзя изменить через массовую загрузку")
                saved = repo.update_product(int(existing["id"]), payload)
                if saved is None:
                    raise ValueError("Товар не найден")
                updated += 1
                old_code_key = str(existing.get("code") or "").strip().casefold()
                existing["code"] = saved.code
                by_code[code_key] = existing
                if old_code_key and old_code_key != code_key:
                    if by_code.get(old_code_key, {}).get("id") == existing["id"]:
                        del by_code[old_code_key]
            else:
                if "barcodes" not in payload:
                    payload["barcodes"] = []
                saved = repo.create_product(payload)
                created += 1
                by_sku[sku_key] = {
                    "id": int(saved.id),
                    "name": saved.name,
                    "sku": saved.sku,
                    "code": saved.code,
                    "is_kit": bool(saved.is_kit),
                    "image_url": saved.image_url or "",
                }
                by_code[code_key] = by_sku[sku_key]

            seen_skus[sku_key] = row_idx
            seen_codes[code_key] = row_idx
            for barcode in barcodes:
                seen_barcodes[barcode.casefold()] = row_idx
        except ValueError as exc:
            failed_rows.append((row_idx, values, str(exc)))

    error_report = _build_error_report(failed_rows) if failed_rows else None
    return CatalogImportResult(
        created=created,
        updated=updated,
        failed=len(failed_rows),
        error_report=error_report,
        total_rows=len(parsed_rows),
    )
