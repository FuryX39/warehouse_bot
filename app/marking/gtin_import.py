"""Массовое внесение GTIN в каталог из Excel."""

from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.catalog_repository import CatalogRepository

_TEMPLATE_HEADERS = ("Артикул", "Код", "Название товара", "GTIN*")
_EXAMPLE_ROW = ("ART-001", "00001", "Пример товара", "4600605012345")
_ERROR_HEADER = "Ошибка"
_ROW_HEADER = "Строка в файле"


@dataclass(frozen=True)
class GtinImportResult:
    created: int
    skipped: int
    failed: int
    error_report: bytes | None
    total_rows: int


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value).strip()
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def build_gtin_import_template(catalog_repo: CatalogRepository) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "GTIN"
    ws.append(list(_TEMPLATE_HEADERS))

    products = catalog_repo.list_product_gtin_rows()
    has_rows = False
    for product in products:
        gtins = product.get("gtins") or []
        if gtins:
            for gtin in gtins:
                ws.append(
                    [
                        product.get("sku") or "",
                        product.get("code") or "",
                        product.get("name") or "",
                        gtin,
                    ]
                )
                has_rows = True
        else:
            ws.append(
                [
                    product.get("sku") or "",
                    product.get("code") or "",
                    product.get("name") or "",
                    "",
                ]
            )
            has_rows = True

    if not has_rows:
        ws.append(list(_EXAMPLE_ROW))

    for col in range(1, len(_TEMPLATE_HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8EEF4")

    ref = wb.create_sheet("Подсказки")
    ref.append(["Поля"])
    ref.append(["Укажите артикул или код товара и GTIN (8, 12, 13 или 14 цифр)."])
    ref.append(["Один товар — несколько строк, если GTIN несколько."])
    ref.append(["Если такой GTIN у товара уже есть, строка пропускается."])
    ref.append(["Колонка «Название товара» только для удобства, при загрузке не используется."])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _row_is_empty(values: list[str]) -> bool:
    return not any((values[0].strip(), values[1].strip(), values[3].strip()))


def _parse_data_rows(sheet) -> list[tuple[int, list[str]]]:
    rows: list[tuple[int, list[str]]] = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = [_cell_str(cell) for cell in row[: len(_TEMPLATE_HEADERS)]]
        while len(values) < len(_TEMPLATE_HEADERS):
            values.append("")
        if _row_is_empty(values):
            continue
        if (
            row_idx == 2
            and values[0].strip().upper() == "ART-001"
            and values[3].strip() == "4600605012345"
        ):
            continue
        rows.append((row_idx, values))
    return rows


def _build_error_report(failed_rows: list[tuple[int, list[str], str]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ошибки"
    headers = [_ROW_HEADER, *_TEMPLATE_HEADERS, _ERROR_HEADER]
    ws.append(headers)
    for row_idx, values, error in failed_rows:
        ws.append([row_idx, *values, error])
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
    fill = PatternFill("solid", fgColor="FFEBEE")
    error_col = len(headers)
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=error_col).fill = fill
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _resolve_product(
    by_sku: dict[str, dict[str, Any]],
    by_code: dict[str, dict[str, Any]],
    *,
    sku: str,
    code: str,
) -> dict[str, Any]:
    sku = sku.strip()
    code = code.strip()
    if not sku and not code:
        raise ValueError("Укажите артикул или код товара")
    found: dict[int, dict[str, Any]] = {}
    if sku:
        product = by_sku.get(sku.casefold())
        if product is None:
            raise ValueError(f"Товар с артикулом «{sku}» не найден")
        found[int(product["id"])] = product
    if code:
        product = by_code.get(code.casefold())
        if product is None:
            raise ValueError(f"Товар с кодом «{code}» не найден")
        found[int(product["id"])] = product
    if len(found) > 1:
        raise ValueError("Артикул и код указывают на разные товары")
    return next(iter(found.values()))


def import_gtins_from_xlsx(catalog_repo: CatalogRepository, data: bytes) -> GtinImportResult:
    if not data:
        raise ValueError("Файл пустой")
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Не удалось прочитать Excel-файл (.xlsx)") from exc

    sheet = wb.active
    if sheet is None:
        raise ValueError("В файле нет листа с данными")

    parsed_rows = _parse_data_rows(sheet)
    data_rows = [item for item in parsed_rows if str(item[1][3]).strip()]
    if not data_rows:
        raise ValueError("В файле нет строк с GTIN для загрузки")

    by_sku, by_code, _by_barcode = catalog_repo.build_product_import_index()
    created = 0
    skipped = 0
    failed_rows: list[tuple[int, list[str], str]] = []

    for row_idx, values in data_rows:
        sku, code, _name, gtin = values
        try:
            if not str(gtin).strip():
                continue
            product = _resolve_product(by_sku, by_code, sku=sku, code=code)
            action = catalog_repo.add_product_gtin(int(product["id"]), gtin)
            if action == "created":
                created += 1
            else:
                skipped += 1
        except ValueError as exc:
            failed_rows.append((row_idx, values, str(exc)))

    error_report = _build_error_report(failed_rows) if failed_rows else None
    return GtinImportResult(
        created=created,
        skipped=skipped,
        failed=len(failed_rows),
        error_report=error_report,
        total_rows=len(data_rows),
    )
