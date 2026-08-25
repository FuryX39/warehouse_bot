"""Excel: артикул — GTIN — коды Data Matrix."""

from __future__ import annotations

import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.marking.cis import replace_gs_for_excel
from app.marking.match import MatchResult, ProductCodeGroup

_CELL_LIMIT = 32000
_CODES_JOIN = "\n"


def _safe(value: object) -> str:
    text = replace_gs_for_excel(str(value or ""))
    return "".join(ch for ch in text if ord(ch) >= 32 or ch in "\t\n\r")


def _style_header(ws: Worksheet, row: int, columns: int) -> None:
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E8EEF4")
    thin = Border(
        left=Side(style="thin", color="D8DEE9"),
        right=Side(style="thin", color="D8DEE9"),
        top=Side(style="thin", color="D8DEE9"),
        bottom=Side(style="thin", color="D8DEE9"),
    )
    for col in range(1, columns + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin


def _append_code_chunks(gtin_codes: list[str]) -> list[str]:
    if not gtin_codes:
        return [""]
    chunks: list[str] = []
    current: list[str] = []
    size = 0
    for code in gtin_codes:
        piece = _safe(code)
        extra = len(piece) + (len(_CODES_JOIN) if current else 0)
        if current and size + extra > _CELL_LIMIT:
            chunks.append(_CODES_JOIN.join(current))
            current = [piece]
            size = len(piece)
        else:
            current.append(piece)
            size += extra
    if current:
        chunks.append(_CODES_JOIN.join(current))
    return chunks


def _write_groups_sheet(ws: Worksheet, groups: list[ProductCodeGroup]) -> None:
    headers = ("Артикул", "GTIN", "Коды DataMatrix")
    ws.append(list(headers))
    _style_header(ws, 1, len(headers))
    wrap = Alignment(wrap_text=True, vertical="top")
    for group in groups:
        chunks = _append_code_chunks(group.codes)
        first = True
        for chunk in chunks:
            ws.append(
                [
                    group.product.sku if first else "",
                    group.gtin if first else "",
                    chunk,
                ]
            )
            ws.cell(row=ws.max_row, column=3).alignment = wrap
            first = False
    ws.column_dimensions[get_column_letter(1)].width = 24
    ws.column_dimensions[get_column_letter(2)].width = 18
    ws.column_dimensions[get_column_letter(3)].width = 80
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:C{max(ws.max_row, 1)}"


def _write_codes_sheet(ws: Worksheet, groups: list[ProductCodeGroup]) -> None:
    headers = ("Артикул", "GTIN", "DataMatrix")
    ws.append(list(headers))
    _style_header(ws, 1, len(headers))
    for group in groups:
        for code in group.codes:
            ws.append([group.product.sku, group.gtin, _safe(code)])
    ws.column_dimensions[get_column_letter(1)].width = 24
    ws.column_dimensions[get_column_letter(2)].width = 18
    ws.column_dimensions[get_column_letter(3)].width = 80
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:C{max(ws.max_row, 1)}"


def _write_other_sheet(ws: Worksheet, headers: tuple[str, ...], rows: list[list[str]]) -> None:
    ws.append(list(headers))
    _style_header(ws, 1, len(headers))
    for row in rows:
        ws.append([_safe(v) for v in row])
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 28 if col < len(headers) else 80
    ws.freeze_panes = "A2"


def build_marking_codes_export(result: MatchResult) -> bytes:
    wb = Workbook()
    summary = wb.active
    summary.title = "Сводка"
    generated = datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d %H:%M")
    summary.append(["Сформировано", generated])
    summary.append(["Строк на входе", result.total_lines])
    summary.append(["Уникальных DataMatrix", result.unique_codes])
    summary.append(["Сопоставлено с товаром", result.matched_code_count])
    summary.append(["Товаров", len(result.groups)])
    summary.append(["GTIN без товара", sum(len(u.codes) for u in result.unmatched)])
    summary.append(["Конфликты GTIN", sum(len(c.codes) for c in result.conflicts)])
    summary.append(["Не разобрано", len(result.invalid)])
    summary.append(["Дубликаты (пропущены)", result.duplicate_count])
    summary.append([])
    summary.append(
        [
            "Разделитель GS (ASCII 29) в кодах записан как <GS> — так Excel не ломает файл.",
        ]
    )
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 22

    groups_ws = wb.create_sheet("По товарам")
    _write_groups_sheet(groups_ws, result.groups)

    codes_ws = wb.create_sheet("Коды")
    _write_codes_sheet(codes_ws, result.groups)

    if result.unmatched:
        other = wb.create_sheet("Без товара")
        rows = []
        for item in result.unmatched:
            for code in item.codes:
                rows.append([item.gtin, code])
        _write_other_sheet(other, ("GTIN", "DataMatrix"), rows)

    if result.conflicts:
        other = wb.create_sheet("Конфликты")
        rows = []
        for item in result.conflicts:
            skus = ", ".join(item.skus)
            for code in item.codes:
                rows.append([item.gtin, skus, code])
        _write_other_sheet(other, ("GTIN", "Артикулы", "DataMatrix"), rows)

    if result.invalid:
        other = wb.create_sheet("Ошибки")
        rows = [[rec.raw, rec.error] for rec in result.invalid]
        _write_other_sheet(other, ("Строка", "Причина"), rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
