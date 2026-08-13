"""Анализ продаж: заказы маркетплейса, сгруппированные по артикулу, в выбранном виде цен."""

from __future__ import annotations

import io
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.catalog_repository import CatalogRepository
from app.repositories import InventoryRepository

MARKETPLACES: tuple[dict[str, str], ...] = (
    {"id": "ozon", "title": "Ozon"},
    {"id": "yandex_market", "title": "Яндекс Маркет"},
    {"id": "wildberries", "title": "Wildberries"},
)
MARKETPLACE_IDS = frozenset(item["id"] for item in MARKETPLACES)

_HEADERS = ("Артикул", "Название", "Количество заказов", "Сумма заказов")


@dataclass(frozen=True)
class SalesAnalysisRow:
    sku: str
    name: str
    quantity: int
    amount: Decimal | None


@dataclass(frozen=True)
class SalesAnalysisResult:
    marketplace_id: str
    marketplace_title: str
    price_type_id: int
    price_type_name: str
    rows: list[SalesAnalysisRow]
    total_quantity: int
    total_amount: Decimal
    missing_price_count: int
    workbook_bytes: bytes


def marketplace_title(marketplace_id: str) -> str:
    for item in MARKETPLACES:
        if item["id"] == marketplace_id:
            return item["title"]
    return marketplace_id


def _excel_safe_text(value: object) -> str:
    text = str(value or "")
    return "".join(ch for ch in text if ord(ch) >= 32 or ch in "\t\n\r")


def _parse_amount(value: object) -> Decimal | None:
    raw = str(value or "").strip().replace("\u00a0", "").replace(" ", "").replace(",", ".")
    if not raw:
        return None
    try:
        amount = Decimal(raw)
    except InvalidOperation:
        return None
    if amount < 0:
        return None
    return amount


def build_sales_analysis(
    inventory_repo: InventoryRepository,
    catalog_repo: CatalogRepository,
    *,
    marketplace_id: str,
    price_type_id: int,
    price_type_name: str,
) -> SalesAnalysisResult:
    source = str(marketplace_id or "").strip()
    if source not in MARKETPLACE_IDS:
        raise ValueError(
            "Маркетплейс должен быть одним из: " + ", ".join(item["title"] for item in MARKETPLACES)
        )
    pt_id = int(price_type_id)
    if pt_id < 1:
        raise ValueError("Некорректный вид цен")
    pt_name = str(price_type_name or "").strip() or f"price_type_{pt_id}"
    title = marketplace_title(source)

    aggregated = inventory_repo.aggregate_order_quantities_by_sku(source)
    catalog_by_sku = catalog_repo.lookup_products_by_skus([sku for sku, _qty in aggregated])
    product_ids = [
        int(item["id"])
        for item in catalog_by_sku.values()
        if item.get("id") is not None
    ]
    prices_by_product = catalog_repo.get_prices_for_products(product_ids, pt_id) if product_ids else {}

    missing_skus = [
        sku for sku, _qty in aggregated if sku.strip().casefold() not in catalog_by_sku
    ]
    nomenclature = inventory_repo.get_nomenclature_meta_for_skus(missing_skus) if missing_skus else {}

    rows: list[SalesAnalysisRow] = []
    total_qty = 0
    total_amount = Decimal("0.00")
    missing_price_count = 0
    for sku, qty in aggregated:
        catalog = catalog_by_sku.get(sku.strip().casefold())
        name = ""
        amount: Decimal | None = None
        if catalog:
            name = str(catalog.get("name") or "").strip()
            product_id = int(catalog["id"])
            unit = _parse_amount(prices_by_product.get(product_id))
            if unit is None:
                missing_price_count += 1
            else:
                amount = (unit * Decimal(qty)).quantize(Decimal("0.01"))
                total_amount += amount
        else:
            meta = nomenclature.get(sku) or {}
            name = str(meta.get("name") or "").strip()
            missing_price_count += 1
        if not name:
            name = "Товар не найден в каталоге"
        total_qty += qty
        rows.append(SalesAnalysisRow(sku=sku, name=name, quantity=qty, amount=amount))

    workbook_bytes = _build_workbook(
        rows,
        marketplace_title=title,
        price_type_name=pt_name,
        total_quantity=total_qty,
        total_amount=total_amount,
    )
    return SalesAnalysisResult(
        marketplace_id=source,
        marketplace_title=title,
        price_type_id=pt_id,
        price_type_name=pt_name,
        rows=rows,
        total_quantity=total_qty,
        total_amount=total_amount.quantize(Decimal("0.01")),
        missing_price_count=missing_price_count,
        workbook_bytes=workbook_bytes,
    )


def _build_workbook(
    rows: list[SalesAnalysisRow],
    *,
    marketplace_title: str,
    price_type_name: str,
    total_quantity: int,
    total_amount: Decimal,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Анализ продаж"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E8EEF4")
    thin = Border(
        left=Side(style="thin", color="D8DEE9"),
        right=Side(style="thin", color="D8DEE9"),
        top=Side(style="thin", color="D8DEE9"),
        bottom=Side(style="thin", color="D8DEE9"),
    )
    money_align = Alignment(horizontal="right")

    ws.append(["Маркетплейс", _excel_safe_text(marketplace_title)])
    ws.append(["Вид цен", _excel_safe_text(price_type_name)])
    ws.append([])
    ws.append(list(_HEADERS))
    for col in range(1, len(_HEADERS) + 1):
        cell = ws.cell(row=4, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin

    for row in rows:
        amount_val: Any = float(row.amount) if row.amount is not None else None
        ws.append(
            [
                _excel_safe_text(row.sku),
                _excel_safe_text(row.name),
                int(row.quantity),
                amount_val,
            ]
        )
        excel_row = ws.max_row
        ws.cell(row=excel_row, column=4).number_format = "#,##0.00"
        ws.cell(row=excel_row, column=4).alignment = money_align
        for col in range(1, 5):
            ws.cell(row=excel_row, column=col).border = thin

    if rows:
        ws.append(["Итого", "", int(total_quantity), float(total_amount)])
        total_row = ws.max_row
        for col in range(1, 5):
            cell = ws.cell(row=total_row, column=col)
            cell.font = header_font
            cell.border = thin
        ws.cell(row=total_row, column=4).number_format = "#,##0.00"
        ws.cell(row=total_row, column=4).alignment = money_align

    ws.column_dimensions[get_column_letter(1)].width = 22
    ws.column_dimensions[get_column_letter(2)].width = 48
    ws.column_dimensions[get_column_letter(3)].width = 22
    ws.column_dimensions[get_column_letter(4)].width = 20
    ws.cell(row=1, column=1).font = header_font

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
