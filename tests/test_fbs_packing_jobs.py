"""Задания FBS-упаковки Яндекс: создание, пик, остаток, close/cancel."""

from __future__ import annotations

import base64
import io
import zipfile
from types import SimpleNamespace
from unittest.mock import patch

from fastapi import FastAPI
from fastapi.testclient import TestClient
from reportlab.pdfgen import canvas

from app.adapters.yandex_market import YandexFbsItem, YandexFbsOrder
from app.catalog_repository import CatalogRepository
from app.config import Settings
from app.crm_repository import CrmRepository
from app.fbs_packing_repository import (
    FbsPackingRepository,
    LINE_DONE,
    LINE_PENDING,
    LINE_PRINTED,
)
from app.fbs_packing_service import create_yandex_packing_job
from app.warehouse_users_repository import WarehouseUserRow
from app.web.warehouse_fbs_packing_routes import register_warehouse_fbs_packing_routes
from app.web.warehouse_tasks_api_auth import TasksApiActor
from app.yandex_fbs_labels import (
    YandexFbsListRow,
    collect_ready_unit_labels,
    collect_started_unit_labels,
)


def _pdf(text: str) -> bytes:
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=(200, 200))
    c.drawString(20, 100, text)
    c.showPage()
    c.save()
    return buf.getvalue()


def _settings(**kwargs) -> Settings:
    payload = {
        "telegram_bot_token": "t",
        "db_url": "sqlite:///:memory:",
        "movement_db_url": "sqlite:///:memory:",
        "yandex_label_format": "A9_HORIZONTALLY",
        "yandex_label_rotate_degrees": 0,
    }
    payload.update(kwargs)
    return Settings(**payload)


def _catalog(tmp_path, name: str = "pack.db") -> tuple[CatalogRepository, str]:
    db_url = f"sqlite:///{(tmp_path / name).as_posix()}"
    crm = CrmRepository(db_url)
    crm.init_schema()
    repo = CatalogRepository(db_url)
    repo.init_schema()
    return repo, db_url


def _packing_repo(tmp_path, db_url: str) -> FbsPackingRepository:
    repo = FbsPackingRepository(db_url, files_data_dir=tmp_path / "fbs_packing")
    repo.init_schema()
    return repo


def _user(user_id: int, login: str) -> WarehouseUserRow:
    return WarehouseUserRow(
        id=user_id,
        login=login,
        display_name=login,
        group_id=None,
        group_name="",
        telegram_nick="",
        is_admin=False,
        is_active=True,
        permissions={},
        created_at_ts=0,
        updated_at_ts=0,
    )


def _order(*, order_id: str, sku: str, quantity: int, item_id: int) -> YandexFbsOrder:
    return YandexFbsOrder(
        order_id=order_id,
        status="PROCESSING",
        substatus="STARTED",
        lines=((sku, quantity),),
        items=(YandexFbsItem(item_id=item_id, sku=sku, quantity=quantity),),
    )


class FakeYandexAdapter:
    def __init__(self, orders: list[YandexFbsOrder]) -> None:
        self.orders = orders
        self.box_calls: list[str] = []
        self._next_box = 7001

    def list_awaiting_assembly_orders(self, *, substatus: str):
        return [order for order in self.orders if order.substatus == substatus]

    def set_order_unit_boxes(self, order: YandexFbsOrder) -> list[int]:
        self.box_calls.append(order.order_id)
        count = sum(max(0, int(item.quantity)) for item in order.items)
        ids = list(range(self._next_box, self._next_box + count))
        self._next_box += count
        return ids

    def fetch_box_label_pdf(self, order_id: str, box_id: int, *, label_format: str) -> bytes:
        assert label_format == "A9_HORIZONTALLY"
        return _pdf(f"{order_id}-{box_id}")

    def fetch_order_label_pdf_parts(self, order_ids, *, label_format):
        raise AssertionError("STARTED must not fetch existing order labels")


def test_create_job_one_unit_one_line_with_pdf(tmp_path) -> None:
    catalog, db_url = _catalog(tmp_path)
    product = catalog.create_product(
        {
            "name": "Болт",
            "sku": "SKU-A",
            "code": "00001",
            "is_kit": False,
            "barcodes": [{"barcode": "2000000000016", "label": "", "group": ""}],
            "components": [],
        }
    )
    packing = _packing_repo(tmp_path, db_url)
    adapter = FakeYandexAdapter(
        [_order(order_id="100001", sku="SKU-A", quantity=2, item_id=501)]
    )

    job = create_yandex_packing_job(
        adapter=adapter,
        catalog=catalog,
        packing_repo=packing,
        settings=_settings(),
        order_substatus="STARTED",
        build_list=False,
        item_limit=None,
        packer_user_ids=[7],
        created_by_user_id=1,
    )

    assert adapter.box_calls == ["100001"]
    assert job.line_total == 2
    assert [line.sku for line in job.lines] == ["SKU-A", "SKU-A"]
    assert [line.product_id for line in job.lines] == [product.id, product.id]
    assert all(line.status == LINE_PENDING for line in job.lines)
    assert all(packing.read_line_pdf(job.id, line.id).startswith(b"%PDF") for line in job.lines)
    assert job.merged_label_stored_name == ""
    assert job.build_list is False


def test_build_list_true_calls_assembly_and_google(tmp_path) -> None:
    catalog, db_url = _catalog(tmp_path)
    packing = _packing_repo(tmp_path, db_url)
    adapter = FakeYandexAdapter(
        [_order(order_id="100001", sku="SKU-MISS", quantity=1, item_id=501)]
    )
    settings = _settings(
        default_stocks_sheet_url="https://docs.google.com/spreadsheets/d/x/edit",
        google_service_account_file="/tmp/creds.json",
        fbs_list_sheet_url="https://docs.google.com/spreadsheets/d/list/edit",
    )

    with (
        patch(
            "app.yandex_fbs_labels.apply_assembly_order_to_yandex_rows",
            side_effect=lambda rows, **_kw: (rows, ["assembly-ok"]),
        ) as assembly,
        patch(
            "app.fbs_packing_service._export_list_to_google_sheet",
            return_value="https://docs.google.com/spreadsheets/d/list/edit#gid=1",
        ) as export,
    ):
        job = create_yandex_packing_job(
            adapter=adapter,
            catalog=catalog,
            packing_repo=packing,
            settings=settings,
            order_substatus="STARTED",
            build_list=True,
            item_limit=None,
            packer_user_ids=[7],
            created_by_user_id=1,
        )

    assembly.assert_called_once()
    export.assert_called_once()
    assert job.build_list is True
    assert job.sheet_url.endswith("gid=1")
    assert job.merged_label_stored_name
    assert packing.read_merged_pdf(job.id).startswith(b"%PDF")
    assert any("не найден в каталоге" in warning for warning in job.warnings)
    assert job.lines[0].product_id is None


def test_build_list_false_skips_assembly_and_google(tmp_path) -> None:
    catalog, db_url = _catalog(tmp_path)
    packing = _packing_repo(tmp_path, db_url)
    adapter = FakeYandexAdapter(
        [_order(order_id="100001", sku="SKU-A", quantity=1, item_id=501)]
    )
    settings = _settings(
        default_stocks_sheet_url="https://docs.google.com/spreadsheets/d/x/edit",
        google_service_account_file="/tmp/creds.json",
        fbs_list_sheet_url="https://docs.google.com/spreadsheets/d/list/edit",
    )

    with (
        patch(
            "app.yandex_fbs_labels.apply_assembly_order_to_yandex_rows",
            side_effect=lambda rows, **_kw: (rows, ["assembly-ok"]),
        ) as assembly,
        patch(
            "app.fbs_packing_service._export_list_to_google_sheet",
            return_value="https://sheet",
        ) as export,
    ):
        job = create_yandex_packing_job(
            adapter=adapter,
            catalog=catalog,
            packing_repo=packing,
            settings=settings,
            order_substatus="STARTED",
            build_list=False,
            item_limit=None,
            packer_user_ids=[7],
            created_by_user_id=1,
        )

    assembly.assert_not_called()
    export.assert_not_called()
    assert job.build_list is False
    assert job.sheet_url == ""
    assert job.merged_label_stored_name == ""


def test_ready_to_ship_skips_one_label_for_many_skus() -> None:
    order = YandexFbsOrder(
        order_id="100001",
        status="PROCESSING",
        substatus="READY_TO_SHIP",
        lines=(("SKU-A", 1), ("SKU-B", 1)),
        items=(
            YandexFbsItem(item_id=501, sku="SKU-A", quantity=1),
            YandexFbsItem(item_id=502, sku="SKU-B", quantity=1),
        ),
    )
    rows = [
        YandexFbsListRow(1, "100001", "SKU-A", 1, "READY_TO_SHIP"),
        YandexFbsListRow(2, "100001", "SKU-B", 1, "READY_TO_SHIP"),
    ]

    class ReadyAdapter:
        def fetch_order_label_pdf_parts(self, order_ids, *, label_format):
            return [("order.pdf", _pdf("100001"))], []

        def set_order_unit_boxes(self, order):
            raise AssertionError("READY_TO_SHIP boxes must not be changed")

    units, warnings = collect_ready_unit_labels(ReadyAdapter(), [order], rows)
    assert units == []
    assert any("один ярлык на несколько товаров" in item for item in warnings)


def test_started_units_follow_list_order() -> None:
    order = YandexFbsOrder(
        order_id="100001",
        status="PROCESSING",
        substatus="STARTED",
        lines=(("SKU-B", 1), ("SKU-A", 1)),
        items=(
            YandexFbsItem(item_id=501, sku="SKU-B", quantity=1),
            YandexFbsItem(item_id=502, sku="SKU-A", quantity=1),
        ),
    )
    rows = [
        YandexFbsListRow(1, "100001", "SKU-A", 1, "STARTED"),
        YandexFbsListRow(2, "100001", "SKU-B", 1, "STARTED"),
    ]
    adapter = FakeYandexAdapter([order])
    units, warnings = collect_started_unit_labels(adapter, [order], rows)
    assert warnings == []
    assert [unit.sku for unit in units] == ["SKU-A", "SKU-B"]
    assert [unit.box_id for unit in units] == [7002, 7001]


def _seed_job(packing: FbsPackingRepository, catalog: CatalogRepository, packer_id: int):
    product_a = catalog.create_product(
        {
            "name": "Товар A",
            "sku": "SKU-A",
            "code": "00001",
            "is_kit": False,
            "barcodes": [{"barcode": "2000000000016", "label": "", "group": ""}],
            "components": [],
        }
    )
    product_b = catalog.create_product(
        {
            "name": "Товар B",
            "sku": "SKU-B",
            "code": "00002",
            "is_kit": False,
            "barcodes": [{"barcode": "2000000000023", "label": "", "group": ""}],
            "components": [],
        }
    )
    job = packing.create_job(
        marketplace="yandex",
        order_substatus="STARTED",
        build_list=False,
        created_by_user_id=1,
        packer_user_ids=[packer_id],
        lines=[
            {
                "seq": 1,
                "sku": "SKU-A",
                "product_id": product_a.id,
                "product_name": "Товар A",
                "order_id": "100001",
                "box_id": 7001,
                "place_index": 1,
                "place_total": 2,
                "scan_keys": ["100001", "100001 1/2", "7001"],
                "pdf": _pdf("A1"),
            },
            {
                "seq": 2,
                "sku": "SKU-A",
                "product_id": product_a.id,
                "product_name": "Товар A",
                "order_id": "100001",
                "box_id": 7002,
                "place_index": 2,
                "place_total": 2,
                "scan_keys": ["100001", "100001 2/2", "7002"],
                "pdf": _pdf("A2"),
            },
            {
                "seq": 3,
                "sku": "SKU-B",
                "product_id": product_b.id,
                "product_name": "Товар B",
                "order_id": "100002",
                "box_id": 7003,
                "place_index": 1,
                "place_total": 1,
                "scan_keys": ["100002", "7003"],
                "pdf": _pdf("B1"),
            },
        ],
    )
    return job, product_a, product_b


def _packing_client(tmp_path):
    catalog, db_url = _catalog(tmp_path, "http.db")
    packing = _packing_repo(tmp_path, db_url)
    packer = _user(7, "packer")
    other = _user(8, "other")
    state = {"user": packer, "actor": TasksApiActor(user=packer, via_api_token=False)}
    job, product_a, product_b = _seed_job(packing, catalog, packer.id)

    app = FastAPI()

    def require_fbs_access():
        return packer

    def require_warehouse_user():
        return state["user"]

    def require_tasks_access():
        return state["actor"]

    register_warehouse_fbs_packing_routes(
        app,
        packing,
        catalog,
        SimpleNamespace(list_assignee_picker=lambda: []),
        _settings(),
        SimpleNamespace(adapters=[]),
        require_fbs_access,
        require_warehouse_user,
        require_tasks_access,
    )
    return TestClient(app), packing, job, state, packer, other, product_a, product_b


def test_packer_scan_pick_remaining_close_cancel(tmp_path) -> None:
    client, packing, job, state, packer, other, product_a, product_b = _packing_client(tmp_path)
    job_id = job.id
    prefix = f"/api/warehouse/fbs-packing/jobs/{job_id}"

    packed = client.get(f"{prefix}/pack")
    assert packed.status_code == 200, packed.text
    remaining = packed.json()["job"]["remaining_groups"]
    assert [(item["sku"], item["quantity"], item["barcode"]) for item in remaining] == [
        ("SKU-A", 2, "2000000000016"),
        ("SKU-B", 1, "2000000000023"),
    ]

    first = client.post(f"{prefix}/scan-product", json={"barcode": "2000000000016"})
    assert first.status_code == 200, first.text
    payload = first.json()
    assert payload["line"]["sku"] == "SKU-A"
    assert payload["line"]["status"] == LINE_PRINTED
    assert payload["line"]["id"] == job.lines[0].id
    assert base64.b64decode(payload["pdf_base64"]).startswith(b"%PDF")
    remaining_after = payload["job"]["remaining_groups"]
    assert [(item["sku"], item["quantity"]) for item in remaining_after] == [
        ("SKU-A", 1),
        ("SKU-B", 1),
    ]

    blocked = client.post(f"{prefix}/scan-product", json={"barcode": "2000000000023"})
    assert blocked.status_code == 400
    assert packing.get_line(job_id, job.lines[0].id).status == LINE_PRINTED

    reprint = client.post(f"{prefix}/scan-product", json={"barcode": "2000000000016"})
    assert reprint.status_code == 200
    assert reprint.json()["line"]["id"] == job.lines[0].id

    mismatch = client.post(f"{prefix}/scan-label", json={"barcode": "100002"})
    assert mismatch.status_code == 400
    assert packing.get_line(job_id, job.lines[0].id).status == LINE_PRINTED

    closed = client.post(f"{prefix}/lines/{job.lines[0].id}/close")
    assert closed.status_code == 200, closed.text
    assert closed.json()["line"]["status"] == LINE_DONE

    second = client.post(f"{prefix}/scan-product", json={"barcode": "2000000000016"})
    assert second.status_code == 200
    assert second.json()["line"]["id"] == job.lines[1].id

    cancelled = client.post(f"{prefix}/lines/{job.lines[1].id}/cancel-print")
    assert cancelled.status_code == 200
    assert cancelled.json()["line"]["status"] == LINE_PENDING

    picked = client.post(
        f"{prefix}/pick-sku",
        json={"sku": "SKU-B", "product_id": product_b.id},
    )
    assert picked.status_code == 200, picked.text
    assert picked.json()["line"]["sku"] == "SKU-B"
    assert picked.json()["line"]["id"] == job.lines[2].id

    labeled = client.post(f"{prefix}/scan-label", json={"barcode": "100002"})
    assert labeled.status_code == 200, labeled.text
    assert labeled.json()["line"]["status"] == LINE_DONE

    label_pdf = client.get(f"{prefix}/lines/{job.lines[2].id}/label")
    assert label_pdf.status_code == 200
    assert label_pdf.content.startswith(b"%PDF")

    state["user"] = other
    forbidden = client.post(f"{prefix}/scan-product", json={"barcode": "2000000000016"})
    assert forbidden.status_code == 403

    state["user"] = packer
    mine = client.get("/api/warehouse/fbs-packing/my")
    assert mine.status_code == 200
    assert mine.json()["jobs"][0]["id"] == job_id

    state["actor"] = TasksApiActor(user=None, via_api_token=True)
    token_only = client.get("/api/v1/fbs-packing/my")
    assert token_only.status_code == 400


def test_packer_downloads_all_line_labels_zip_and_can_skip_pdf(tmp_path) -> None:
    client, packing, job, state, packer, other, product_a, product_b = _packing_client(tmp_path)
    job_id = job.id
    prefix = f"/api/warehouse/fbs-packing/jobs/{job_id}"

    zipped = client.get(f"{prefix}/line-labels.zip")
    assert zipped.status_code == 200, zipped.text
    assert zipped.headers["content-type"].startswith("application/zip")
    with zipfile.ZipFile(io.BytesIO(zipped.content)) as archive:
        names = set(archive.namelist())
        assert names == {f"{line.id}.pdf" for line in job.lines}
        for line in job.lines:
            assert archive.read(f"{line.id}.pdf").startswith(b"%PDF")

    skipped = client.post(
        f"{prefix}/scan-product",
        json={"barcode": "2000000000016", "include_pdf": False},
    )
    assert skipped.status_code == 200, skipped.text
    payload = skipped.json()
    assert payload["pdf_base64"] == ""
    assert payload["pdfs_base64"] == []
    assert payload["line"]["id"] == job.lines[0].id


def test_packer_auto_close_marks_line_done_and_frees_next_sku(tmp_path) -> None:
    client, packing, job, state, packer, other, product_a, product_b = _packing_client(tmp_path)
    job_id = job.id
    prefix = f"/api/warehouse/fbs-packing/jobs/{job_id}"

    first = client.post(
        f"{prefix}/scan-product",
        json={"barcode": "2000000000016", "include_pdf": False, "auto_close": True},
    )
    assert first.status_code == 200, first.text
    payload = first.json()
    assert payload["line"]["status"] == LINE_DONE
    assert payload["pdfs_base64"] == []
    assert packing.get_line(job_id, job.lines[0].id).status == LINE_DONE

    second = client.post(
        f"{prefix}/scan-product",
        json={"barcode": "2000000000023", "include_pdf": False, "auto_close": True},
    )
    assert second.status_code == 200, second.text
    assert second.json()["line"]["id"] == job.lines[2].id
    assert second.json()["line"]["status"] == LINE_DONE


def test_packer_batch_allocate_prints_all_sku_then_scan_labels(tmp_path) -> None:
    client, packing, job, state, packer, other, product_a, product_b = _packing_client(tmp_path)
    job_id = job.id
    prefix = f"/api/warehouse/fbs-packing/jobs/{job_id}"

    batch = client.post(f"{prefix}/scan-product", json={"barcode": "2000000000016", "batch": True})
    assert batch.status_code == 200, batch.text
    payload = batch.json()
    assert len(payload["lines"]) == 2
    assert [line["id"] for line in payload["lines"]] == [job.lines[0].id, job.lines[1].id]
    assert len(payload["pdfs_base64"]) == 2
    assert all(base64.b64decode(item).startswith(b"%PDF") for item in payload["pdfs_base64"])
    assert [(item["sku"], item["quantity"]) for item in payload["job"]["remaining_groups"]] == [
        ("SKU-B", 1),
    ]
    assert len(payload["job"]["active_lines"]) == 2

    blocked = client.post(f"{prefix}/scan-product", json={"barcode": "2000000000023", "batch": True})
    assert blocked.status_code == 400

    first_label = client.post(f"{prefix}/scan-label", json={"barcode": "100001 2/2"})
    assert first_label.status_code == 200, first_label.text
    assert first_label.json()["line"]["id"] == job.lines[1].id
    assert first_label.json()["line"]["status"] == LINE_DONE
    assert len(first_label.json()["job"]["active_lines"]) == 1

    second_label = client.post(f"{prefix}/scan-label", json={"barcode": "100001 1/2"})
    assert second_label.status_code == 200, second_label.text
    assert second_label.json()["line"]["id"] == job.lines[0].id
    assert second_label.json()["job"]["active_lines"] == []

    again = client.post(
        f"{prefix}/pick-sku",
        json={"sku": "SKU-B", "product_id": product_b.id, "batch": True},
    )
    assert again.status_code == 200, again.text
    assert len(again.json()["lines"]) == 1
    assert again.json()["line"]["sku"] == "SKU-B"


def test_packer_batch_cancel_resets_all_printed(tmp_path) -> None:
    client, packing, job, *_rest = _packing_client(tmp_path)
    job_id = job.id
    prefix = f"/api/warehouse/fbs-packing/jobs/{job_id}"

    batch = client.post(f"{prefix}/scan-product", json={"barcode": "2000000000016", "batch": 1})
    assert batch.status_code == 200, batch.text
    line_id = batch.json()["lines"][0]["id"]

    cancelled = client.post(f"{prefix}/lines/{line_id}/cancel-print")
    assert cancelled.status_code == 200, cancelled.text
    assert packing.get_line(job_id, job.lines[0].id).status == LINE_PENDING
    assert packing.get_line(job_id, job.lines[1].id).status == LINE_PENDING
    assert packing.get_line(job_id, job.lines[2].id).status == LINE_PENDING


GTIN13 = "4600605012345"
GTIN14 = "04600605012345"


def _cis(gtin14: str, serial: str = "Ab12Xy", *, with_gs: bool = True) -> str:
    from app.marking.cis import GS

    body = f"01{gtin14}21{serial}"
    crypto = "91FFD092abcdef0123456789"
    if with_gs:
        return f"{body}{GS}91FFD0{GS}92abcdef0123456789"
    return body + crypto


def _seed_cis_job(packing: FbsPackingRepository, catalog: CatalogRepository, packer_id: int, *, require_cis: bool):
    product = catalog.create_product(
        {
            "name": "Маркируемый",
            "sku": "SKU-CIS",
            "code": "00099",
            "is_kit": False,
            "barcodes": [{"barcode": "2000000000092", "label": "", "group": ""}],
            "gtins": [GTIN13],
            "components": [],
        }
    )
    plain = catalog.create_product(
        {
            "name": "Без маркировки",
            "sku": "SKU-PLAIN",
            "code": "00098",
            "is_kit": False,
            "barcodes": [{"barcode": "PLAIN-NO-GTIN", "label": "", "group": ""}],
            "components": [],
        }
    )
    job = packing.create_job(
        marketplace="yandex",
        order_substatus="STARTED",
        build_list=False,
        require_cis=require_cis,
        created_by_user_id=1,
        packer_user_ids=[packer_id],
        lines=[
            {
                "seq": 1,
                "sku": "SKU-CIS",
                "product_id": product.id,
                "product_name": "Маркируемый",
                "order_id": "200001",
                "box_id": 8001,
                "place_index": 1,
                "place_total": 2,
                "scan_keys": ["200001", "200001 1/2", "8001"],
                "pdf": _pdf("C1"),
            },
            {
                "seq": 2,
                "sku": "SKU-CIS",
                "product_id": product.id,
                "product_name": "Маркируемый",
                "order_id": "200001",
                "box_id": 8002,
                "place_index": 2,
                "place_total": 2,
                "scan_keys": ["200001", "200001 2/2", "8002"],
                "pdf": _pdf("C2"),
            },
            {
                "seq": 3,
                "sku": "SKU-PLAIN",
                "product_id": plain.id,
                "product_name": "Без маркировки",
                "order_id": "200002",
                "box_id": 8003,
                "place_index": 1,
                "place_total": 1,
                "scan_keys": ["200002", "8003"],
                "pdf": _pdf("P1"),
            },
        ],
    )
    return job, product, plain


def _cis_client(tmp_path, *, require_cis: bool = False):
    catalog, db_url = _catalog(tmp_path, "cis.db")
    packing = _packing_repo(tmp_path, db_url)
    packer = _user(7, "packer")
    job, product, plain = _seed_cis_job(packing, catalog, packer.id, require_cis=require_cis)

    app = FastAPI()

    def require_fbs_access():
        return packer

    def require_warehouse_user():
        return packer

    def require_tasks_access():
        return TasksApiActor(user=packer, via_api_token=False)

    register_warehouse_fbs_packing_routes(
        app,
        packing,
        catalog,
        SimpleNamespace(list_assignee_picker=lambda: []),
        _settings(),
        SimpleNamespace(adapters=[]),
        require_fbs_access,
        require_warehouse_user,
        require_tasks_access,
    )
    return TestClient(app), packing, job, product, plain


def test_cis_scan_allocates_one_line_even_with_batch(tmp_path) -> None:
    client, packing, job, product, _plain = _cis_client(tmp_path, require_cis=False)
    prefix = f"/api/warehouse/fbs-packing/jobs/{job.id}"
    cis = _cis(GTIN14, serial="Ser001")

    resp = client.post(f"{prefix}/scan-product", json={"barcode": cis, "batch": True})
    assert resp.status_code == 200, resp.text
    payload = resp.json()
    assert len(payload["lines"]) == 1
    assert len(payload["pdfs_base64"]) == 1
    assert payload["line"]["sku"] == "SKU-CIS"
    assert payload["line"]["has_cis"] is True
    line = packing.get_line(job.id, payload["line"]["id"])
    assert line is not None
    assert line.cis_gtin == GTIN14
    assert line.cis_key
    assert packing.get_line(job.id, job.lines[1].id).status == LINE_PENDING


def test_require_cis_rejects_plain_barcode_for_gtin_product(tmp_path) -> None:
    client, packing, job, product, _plain = _cis_client(tmp_path, require_cis=True)
    prefix = f"/api/warehouse/fbs-packing/jobs/{job.id}"

    rejected = client.post(f"{prefix}/scan-product", json={"barcode": "2000000000092"})
    assert rejected.status_code == 400
    assert "КИЗ" in rejected.json()["detail"]

    ok = client.post(f"{prefix}/scan-product", json={"barcode": _cis(GTIN14, serial="SerOk1")})
    assert ok.status_code == 200, ok.text
    assert ok.json()["line"]["has_cis"] is True


def test_require_cis_off_allows_barcode_or_cis(tmp_path) -> None:
    client, packing, job, product, _plain = _cis_client(tmp_path, require_cis=False)
    prefix = f"/api/warehouse/fbs-packing/jobs/{job.id}"

    by_barcode = client.post(f"{prefix}/scan-product", json={"barcode": "2000000000092"})
    assert by_barcode.status_code == 200, by_barcode.text
    assert by_barcode.json()["line"]["has_cis"] is False
    line_id = by_barcode.json()["line"]["id"]
    client.post(f"{prefix}/lines/{line_id}/cancel-print")

    by_cis = client.post(f"{prefix}/scan-product", json={"barcode": _cis(GTIN14, serial="SerAlt")})
    assert by_cis.status_code == 200, by_cis.text
    assert by_cis.json()["line"]["has_cis"] is True


def test_duplicate_cis_rejected(tmp_path) -> None:
    client, packing, job, product, _plain = _cis_client(tmp_path, require_cis=False)
    prefix = f"/api/warehouse/fbs-packing/jobs/{job.id}"
    cis = _cis(GTIN14, serial="DupSer")

    first = client.post(f"{prefix}/scan-product", json={"barcode": cis})
    assert first.status_code == 200, first.text
    line_id = first.json()["line"]["id"]
    labeled = client.post(f"{prefix}/scan-label", json={"barcode": "200001 1/2"})
    assert labeled.status_code == 200, labeled.text
    assert packing.get_line(job.id, line_id).status == LINE_DONE

    again = client.post(f"{prefix}/scan-product", json={"barcode": cis})
    assert again.status_code == 400
    assert "уже использован" in again.json()["detail"]


def test_cancel_print_clears_cis(tmp_path) -> None:
    client, packing, job, product, _plain = _cis_client(tmp_path, require_cis=False)
    prefix = f"/api/warehouse/fbs-packing/jobs/{job.id}"
    cis = _cis(GTIN14, serial="CancelMe")

    printed = client.post(f"{prefix}/scan-product", json={"barcode": cis})
    assert printed.status_code == 200, printed.text
    line_id = printed.json()["line"]["id"]
    assert packing.get_line(job.id, line_id).cis_key

    cancelled = client.post(f"{prefix}/lines/{line_id}/cancel-print")
    assert cancelled.status_code == 200, cancelled.text
    line = packing.get_line(job.id, line_id)
    assert line is not None
    assert line.status == LINE_PENDING
    assert line.cis_key == ""
    assert line.cis_raw == ""
    assert line.cis_gtin == ""

    again = client.post(f"{prefix}/scan-product", json={"barcode": cis})
    assert again.status_code == 200, again.text


def test_plain_ean_does_not_write_cis_fields(tmp_path) -> None:
    client, packing, job, product, _plain = _cis_client(tmp_path, require_cis=False)
    prefix = f"/api/warehouse/fbs-packing/jobs/{job.id}"

    resp = client.post(f"{prefix}/scan-product", json={"barcode": "2000000000092"})
    assert resp.status_code == 200, resp.text
    line = packing.get_line(job.id, resp.json()["line"]["id"])
    assert line is not None
    assert line.cis_key == ""
    assert line.cis_gtin == ""
    assert resp.json()["line"]["has_cis"] is False


def test_marking_xlsx_two_sheets(tmp_path) -> None:
    from io import BytesIO

    from openpyxl import load_workbook

    client, packing, job, product, _plain = _cis_client(tmp_path, require_cis=False)
    prefix = f"/api/warehouse/fbs-packing/jobs/{job.id}"

    cis = _cis(GTIN14, serial="Xlsx01")
    printed = client.post(f"{prefix}/scan-product", json={"barcode": cis})
    assert printed.status_code == 200, printed.text
    line_id = printed.json()["line"]["id"]
    done = client.post(f"{prefix}/scan-label", json={"barcode": "200001 1/2"})
    assert done.status_code == 200, done.text

    xlsx = client.get(f"/api/warehouse/fbs-packing/jobs/{job.id}/marking.xlsx")
    assert xlsx.status_code == 200, xlsx.text
    wb = load_workbook(BytesIO(xlsx.content))
    assert wb.sheetnames == ["Все строки", "С КИЗ"]
    full = wb["Все строки"]
    only = wb["С КИЗ"]
    assert full.max_row == 4  # header + 3 lines
    assert only.max_row == 2  # header + 1 with CIS
    assert full.cell(1, 1).value == "Заказ"
    assert full.cell(1, 2).value == "SKU"
    assert full.cell(1, 3).value == "КИЗ"
    # first data row is seq 1 with CIS
    assert full.cell(2, 2).value == "SKU-CIS"
    assert full.cell(2, 3).value
    assert "<GS>" in str(full.cell(2, 3).value) or str(full.cell(2, 3).value).startswith("01")
    assert only.cell(2, 2).value == "SKU-CIS"
    assert packing.get_line(job.id, line_id).status == LINE_DONE


def test_require_cis_blocks_pick_sku_for_markable(tmp_path) -> None:
    client, packing, job, product, plain = _cis_client(tmp_path, require_cis=True)
    prefix = f"/api/warehouse/fbs-packing/jobs/{job.id}"

    blocked = client.post(
        f"{prefix}/pick-sku",
        json={"sku": "SKU-CIS", "product_id": product.id},
    )
    assert blocked.status_code == 400

    allowed = client.post(
        f"{prefix}/pick-sku",
        json={"sku": "SKU-PLAIN", "product_id": plain.id},
    )
    assert allowed.status_code == 200, allowed.text
    assert allowed.json()["line"]["sku"] == "SKU-PLAIN"
