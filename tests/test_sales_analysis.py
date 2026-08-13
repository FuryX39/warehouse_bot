"""Анализ продаж: группировка заказов маркетплейса и Excel."""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.adapters.base import ReservationAction
from app.catalog_repository import CatalogRepository
from app.crm_repository import CrmRepository
from app.repositories import InventoryRepository, OrderItem
from app.sales_analysis import build_sales_analysis


def _setup(tmp_path):
    db_url = f"sqlite:///{(tmp_path / 'sales.db').as_posix()}"
    crm = CrmRepository(db_url)
    crm.init_schema()
    catalog = CatalogRepository(db_url)
    catalog.init_schema()
    inventory = InventoryRepository(db_url)
    inventory.init_schema()
    return crm, catalog, inventory


def test_sales_analysis_sums_orders_by_sku_and_price_type(tmp_path) -> None:
    crm, catalog, inventory = _setup(tmp_path)
    pt = crm.get_meta()["price_types"][0]
    catalog.create_product(
        {
            "name": "Товар А",
            "sku": "SKU-A",
            "code": "00001",
            "is_kit": False,
            "barcodes": [],
            "components": [],
            "prices": [{"price_type_id": pt["id"], "price": "100.00"}],
        }
    )
    catalog.create_product(
        {
            "name": "Товар Б",
            "sku": "SKU-B",
            "code": "00002",
            "is_kit": False,
            "barcodes": [],
            "components": [],
            "prices": [{"price_type_id": pt["id"], "price": "50.50"}],
        }
    )
    inventory.upsert_order_items_from_actions(
        [
            ReservationAction("ozon", "111:SKU-A", "SKU-A", 2),
            ReservationAction("ozon", "222:SKU-A", "SKU-A", 3),
            ReservationAction("ozon", "333:SKU-B", "SKU-B", 1),
            ReservationAction("wildberries", "wb-1", "SKU-A", 10),
        ],
        sync_ts=1_700_000_000,
    )
    with Session(inventory.engine) as session:
        session.add(
            OrderItem(
                source="ozon",
                external_order_id="cancelled:SKU-A",
                sku="SKU-A",
                quantity=99,
                state="cancelled",
                first_seen_ts=1,
                last_seen_ts=1,
            )
        )
        session.commit()

    result = build_sales_analysis(
        inventory,
        catalog,
        marketplace_id="ozon",
        price_type_id=int(pt["id"]),
        price_type_name=str(pt["name"]),
    )
    by_sku = {row.sku: row for row in result.rows}
    assert by_sku["SKU-A"].quantity == 5
    assert str(by_sku["SKU-A"].amount) == "500.00"
    assert by_sku["SKU-A"].name == "Товар А"
    assert by_sku["SKU-B"].quantity == 1
    assert str(by_sku["SKU-B"].amount) == "50.50"
    assert result.total_quantity == 6
    assert str(result.total_amount) == "550.50"
    assert result.missing_price_count == 0

    wb = load_workbook(BytesIO(result.workbook_bytes), data_only=True)
    ws = wb.active
    assert ws["B1"].value == "Ozon"
    assert ws["B2"].value == pt["name"]
    assert ws["A4"].value == "Артикул"
    values = [tuple(row) for row in ws.iter_rows(min_row=5, values_only=True)]
    assert ("SKU-A", "Товар А", 5, 500) in values or ("SKU-A", "Товар А", 5, 500.0) in values
    assert any(row and row[0] == "Итого" and row[2] == 6 for row in values)


def test_sales_analysis_missing_catalog_price_and_unknown_sku(tmp_path) -> None:
    crm, catalog, inventory = _setup(tmp_path)
    pt = crm.get_meta()["price_types"][0]
    catalog.create_product(
        {
            "name": "Без цены",
            "sku": "NO-PRICE",
            "code": "00003",
            "is_kit": False,
            "barcodes": [],
            "components": [],
            "prices": [],
        }
    )
    inventory.upsert_nomenclature_items({"ORPHAN": ("Сирота", "", [])})
    inventory.upsert_order_items_from_actions(
        [
            ReservationAction("yandex_market", "ym-1", "NO-PRICE", 2),
            ReservationAction("yandex_market", "ym-2", "ORPHAN", 4),
        ],
        sync_ts=1_700_000_000,
    )
    result = build_sales_analysis(
        inventory,
        catalog,
        marketplace_id="yandex_market",
        price_type_id=int(pt["id"]),
        price_type_name=str(pt["name"]),
    )
    by_sku = {row.sku: row for row in result.rows}
    assert by_sku["NO-PRICE"].amount is None
    assert by_sku["ORPHAN"].name == "Сирота"
    assert by_sku["ORPHAN"].amount is None
    assert result.missing_price_count == 2
    assert str(result.total_amount) == "0.00"
