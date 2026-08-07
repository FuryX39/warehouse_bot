"""Выгрузка вида цен в Excel."""

from __future__ import annotations

from openpyxl import load_workbook
from starlette.responses import Response

from app.catalog_price_type_import import build_price_type_prices_export
from app.catalog_repository import CatalogRepository
from app.crm_repository import CrmRepository
from app.web.warehouse_catalog_routes import _attachment_disposition


def test_attachment_disposition_allows_cyrillic_filename() -> None:
    header = _attachment_disposition("price_type_Розничная.xlsx")
    response = Response(content=b"x", headers={"Content-Disposition": header})
    # uvicorn кодирует raw_headers в latin-1 — здесь и был 500
    assert response.raw_headers
    assert "filename*=UTF-8''" in header
    assert "Розничная" not in header.split("filename*=", 1)[0]


def test_build_price_type_prices_export_fills_prices(tmp_path) -> None:
    db_url = f"sqlite:///{(tmp_path / 'pt_export.db').as_posix()}"
    crm = CrmRepository(db_url)
    crm.init_schema()
    catalog = CatalogRepository(db_url)
    catalog.init_schema()

    price_types = crm.get_meta()["price_types"]
    assert price_types
    pt = price_types[0]

    product = catalog.create_product(
        {
            "name": "Товар\x00с нулём",
            "sku": "SKU-EXP-1",
            "code": "00099",
            "is_kit": False,
            "barcodes": [{"barcode": "ABC12345", "label": "", "group": ""}],
            "components": [],
            "prices": [{"price_type_id": pt["id"], "price": "199.50"}],
        }
    )
    assert product.id

    raw = build_price_type_prices_export(
        catalog,
        price_type_id=int(pt["id"]),
        price_type_name=str(pt["name"]),
    )
    wb = load_workbook(__import__("io").BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    assert ws["A1"].value.startswith("Вид цен")
    assert ws["B1"].value == pt["name"]
    assert ws["A2"].value == "Артикул"
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    assert any(row and row[0] == "SKU-EXP-1" and str(row[4]) == "199.50" for row in rows)
