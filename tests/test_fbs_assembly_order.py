"""Тесты порядка FBS-списка по листу assembly (формат tables_examples/list.xlsx)."""

from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace

import openpyxl

from app.fbs_assembly_order import (
    parse_assembly_sheet_values,
    reorder_ozon_fbs_list_rows,
    sku_match_key,
)
from app.ozon_fbs_labels import OzonFbsListRow


def _sample_assembly_values() -> list[list[str]]:
    """Минимальный фрагмент листа ТСД: колонки A (артикул), C (ячейка), E (кол-во)."""
    return [
        ["Номенклатура", "", "Ячейка", "", "Количество"],
        ["SSBL997", "", "A-01", "", "1"],
        ["SS414", "", "B-02", "", "2"],
        ["SS996", "", "C-03", "", "1"],
    ]


def test_parse_assembly_sheet_values_reads_fixed_columns_a_c_e() -> None:
    entries = parse_assembly_sheet_values(_sample_assembly_values())
    assert len(entries) == 3
    assert entries[0].sku == "SSBL997"
    assert entries[0].place == "A-01"
    assert entries[0].quantity == 1
    assert entries[0].sort_index == 0
    assert entries[2].sku == "SS996"


def test_parse_uses_column_a_not_other_columns() -> None:
    values = [
        ["Номенклатура", "Артикул", "Ячейка"],
        ["SS278", "SS999", "A-01", "", "3"],
    ]
    entries = parse_assembly_sheet_values(values)
    assert len(entries) == 1
    assert entries[0].sku == "SS278"
    assert entries[0].quantity == 3


def test_reorder_follows_column_a_top_to_bottom() -> None:
    entries = parse_assembly_sheet_values(_sample_assembly_values())
    rows = [
        SimpleNamespace(posting_number="P1", sku="SS996", quantity=1, status=""),
        SimpleNamespace(posting_number="P2", sku="SS414", quantity=2, status=""),
        SimpleNamespace(posting_number="P3", sku="SSBL997", quantity=1, status=""),
    ]
    out = reorder_ozon_fbs_list_rows(rows, entries, row_factory=OzonFbsListRow)
    assert [(r.sku, r.posting_number) for r in out] == [
        ("SSBL997", "P3"),
        ("SS414", "P2"),
        ("SS996", "P1"),
    ]


def test_reorder_same_sku_keeps_api_order_within_sku() -> None:
    entries = parse_assembly_sheet_values(
        [
            ["Номенклатура", "", "Ячейка"],
            ["SS278", "", "A-1"],
            ["SS278", "", "A-2"],
            ["SS533", "", "B-1"],
        ]
    )
    rows = [
        SimpleNamespace(posting_number="P1", sku="SS278", quantity=1, status=""),
        SimpleNamespace(posting_number="P2", sku="SS278", quantity=1, status=""),
        SimpleNamespace(posting_number="P3", sku="SS533", quantity=1, status=""),
    ]
    out = reorder_ozon_fbs_list_rows(rows, entries, row_factory=OzonFbsListRow)
    assert [(r.sku, r.posting_number) for r in out] == [
        ("SS278", "P1"),
        ("SS278", "P2"),
        ("SS533", "P3"),
    ]


def test_sku_match_key_case_insensitive() -> None:
    assert sku_match_key("SS278") == sku_match_key("ss278")
    assert sku_match_key(" SS278 ") == sku_match_key("SS278")
    assert sku_match_key("'SS278") == sku_match_key("SS278")


def test_reorder_real_test_xlsx_by_assembly_column_a() -> None:
    path = Path(__file__).resolve().parents[1] / "tables_examples" / "test.xlsx"
    if not path.exists():
        return

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        assembly_ws = wb["assembly"]
        fbs_ws = wb["FBS"]
        entries = parse_assembly_sheet_values(
            [[cell for cell in row] for row in assembly_ws.iter_rows(values_only=True)]
        )
        rows = [
            SimpleNamespace(
                posting_number=str(row[4]),
                sku=str(row[0]),
                quantity=int(row[3] or 0),
                status="",
            )
            for row in fbs_ws.iter_rows(min_row=2, values_only=True)
            if row[0] and row[4]
        ]
    finally:
        wb.close()

    out = reorder_ozon_fbs_list_rows(rows, entries, row_factory=OzonFbsListRow)
    assert entries[:6] and [e.sku for e in entries[:6]] == [
        "SS821",
        "SS696",
        "SS562",
        "SS940",
        "SS774",
        "SS424",
    ]
    assert [r.sku for r in out[:38]] == (
        ["SS821"] * 23
        + ["SS696"] * 3
        + ["SS562"] * 5
        + ["SS940"]
        + ["SS774"] * 4
        + ["SS424"] * 2
    )
