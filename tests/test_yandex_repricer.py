"""Тесты репрайсера Яндекс Маркет."""

from __future__ import annotations

import io

from openpyxl import Workbook, load_workbook

from app.yandex_repricer import (
    _needs_reprice,
    card_from_seller,
    card_from_showcase,
    process_yandex_prices_workbook,
    seller_from_showcase,
    seller_from_target_card,
    showcase_from_seller,
)


def test_needs_reprice_thresholds() -> None:
    catalog = 1000.0
    assert not _needs_reprice(991, catalog)
    assert _needs_reprice(990, catalog)
    assert not _needs_reprice(1000, catalog)
    assert not _needs_reprice(1099, catalog)
    assert _needs_reprice(1100, catalog)


def test_seller_from_target_card_never_undershoots_on_raise() -> None:
    for target in range(50, 2500, 7):
        recommended = seller_from_target_card(float(target), raise_price=True)
        assert card_from_seller(recommended) >= target


def test_seller_from_target_card_never_overshoots_on_lower() -> None:
    for target in range(50, 2500, 7):
        current = seller_from_target_card(float(target), raise_price=True) + 50
        recommended = seller_from_target_card(
            float(target),
            current_seller=current,
            raise_price=False,
        )
        assert card_from_seller(recommended) <= target


def test_seller_from_target_card_raise_respects_current_seller() -> None:
    recommended = seller_from_target_card(300.0, current_seller=400, raise_price=True)
    assert recommended >= 400
    assert card_from_seller(recommended) >= 300


def test_seller_from_target_card_fixes_known_undershoot() -> None:
    recommended = seller_from_target_card(202.0, raise_price=True)
    assert card_from_seller(recommended) >= 202
    assert recommended == 331

def test_showcase_and_card_formulas() -> None:
    assert showcase_from_seller(300) == 254
    assert card_from_showcase(253) == 182
    assert showcase_from_seller(1239) == 754
    assert card_from_showcase(754) == 626


def test_card_curve_matches_more_prices_examples() -> None:
    examples = [
        (252, 181),
        (345, 259),
        (615, 492),
        (874, 752),
        (922, 820),
        (1257, 1219),
        (2177, 2112),
    ]
    errors = [abs(card_from_showcase(showcase) - actual) for showcase, actual in examples]
    assert sum(errors) / len(errors) <= 2


def test_heavy_coinvest_uses_diminishing_return_when_raising_price() -> None:
    current_seller = 312
    current_showcase = 168
    ratio = current_showcase / current_seller
    target = 190

    linear = seller_from_target_card(
        target,
        current_seller=current_seller,
        raise_price=True,
        showcase_ratio=ratio,
    )
    corrected = seller_from_target_card(
        target,
        current_seller=current_seller,
        current_showcase=current_showcase,
        raise_price=True,
        showcase_ratio=ratio,
    )

    assert linear == 490
    assert corrected > linear


def test_high_showcase_ratio_stays_linear_when_raising_price() -> None:
    current_seller = 1000
    current_showcase = 960
    ratio = current_showcase / current_seller
    target = 1000

    linear = seller_from_target_card(
        target,
        current_seller=current_seller,
        raise_price=True,
        showcase_ratio=ratio,
    )
    corrected = seller_from_target_card(
        target,
        current_seller=current_seller,
        current_showcase=current_showcase,
        raise_price=True,
        showcase_ratio=ratio,
    )

    assert corrected == linear


def test_seller_from_target_card_inverts_chain() -> None:
    target = 599
    recommended = seller_from_target_card(target, raise_price=True)
    assert card_from_seller(recommended) >= target


def test_seller_from_showcase_breakpoint() -> None:
    assert seller_from_showcase(253) < 500
    assert seller_from_showcase(754) >= 500


def test_output_sets_old_price_to_double_recommended_price() -> None:
    class Repo:
        def list_products_for_price_type(self, price_type_id, filters):
            return [{"sku": "SKU-1", "price": 500}]

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "Ваш SKU *",
            "Название товара",
            "Цена *",
            "Зачёркнутая цена",
            "Минимум для акции",
            "На витрине",
        ]
    )
    sheet.append(["SKU-1", "Товар", 1000, 2000, 1000, 900])
    source = io.BytesIO()
    workbook.save(source)

    result = process_yandex_prices_workbook(
        source.getvalue(),
        catalog_repo=Repo(),
        price_type_id=1,
    )

    output = load_workbook(io.BytesIO(result.workbook_bytes), data_only=True).active
    seller_price = output.cell(row=2, column=3).value
    assert output.cell(row=2, column=4).value == seller_price * 2
    assert output.cell(row=2, column=5).value == seller_price
