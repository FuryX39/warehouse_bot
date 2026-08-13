"""Отчёт WB об издержках на приём платежей."""

from __future__ import annotations

from io import BytesIO

from decimal import Decimal

from openpyxl import load_workbook

from app.wb_acquiring_report import (
    AcquiringReportMeta,
    FinanceRateLimiter,
    build_acquiring_workbook,
    build_month_acquiring_report,
    parse_year_month,
)


def test_parse_year_month() -> None:
    assert parse_year_month("2026-04") == (2026, 4)


def test_workbook_net_vat_subtracts_returns() -> None:
    reports = [
        AcquiringReportMeta(
            report_id=731823645,
            seller_name="ШАЙН СИСТЕМС ООО",
            date_from="2026-04-01",
            date_to="2026-04-30",
            create_date="2026-05-01",
            currency="RUB",
            acquiring_fee_sum="921375.53",
            acquiring_fee_vat_sum="40803.84",
        )
    ]
    rows = [
        {
            "rrdId": 1,
            "reportId": 731823645,
            "documentType": "Продажа",
            "retailAmount": "1000",
            "acquiringFee": "100.00",
            "acquiringFeeVat": "20.00",
            "currency": "RUB",
        },
        {
            "rrdId": 2,
            "reportId": 731823645,
            "documentType": "Возврат",
            "retailAmount": "50",
            "acquiringFee": "5.00",
            "acquiringFeeVat": "1.00",
            "currency": "RUB",
        },
    ]
    raw = build_acquiring_workbook(year=2026, month=4, reports=reports, rows=rows)
    wb = load_workbook(BytesIO(raw), data_only=True)
    summary = {str(row[0].value): row[1].value for row in wb["Сводка"].iter_rows(min_row=3, max_row=16)}
    assert summary["№ отчёта"] == 731823645 or summary["№ отчёта"] == "731823645"
    assert float(summary["Сумма издержек по эквайрингу (нетто)"]) == 95.0
    assert float(summary["В том числе НДС (нетто)"]) == 19.0
    details = list(wb["Детализация"].iter_rows(min_row=2, values_only=True))
    assert len(details) == 2
    vat_sum = sum(float(row[12] or 0) for row in details)
    assert vat_sum == 21.0


def test_build_month_acquiring_report_paginates() -> None:
    calls: list[tuple[str, dict]] = []

    def post_fn(url: str, payload: dict):
        calls.append((url, dict(payload)))
        if url.endswith("/acquiring/list"):
            return (
                200,
                [
                    {
                        "reportId": 11,
                        "sellerFinanceName": "ООО Тест",
                        "dateFrom": "2026-04-01",
                        "dateTo": "2026-04-30",
                        "createDate": "2026-05-01",
                        "currency": "RUB",
                        "acquiringFeeSum": "15.00",
                        "acquiringFeeVatSum": "3.00",
                    }
                ],
                "",
            )
        if payload.get("rrdId") == 0:
            return (
                200,
                [
                    {
                        "rrdId": 100,
                        "reportId": 11,
                        "documentType": "Продажа",
                        "acquiringFee": "10",
                        "acquiringFeeVat": "2",
                    },
                    {
                        "rrdId": 101,
                        "reportId": 11,
                        "documentType": "Продажа",
                        "acquiringFee": "8",
                        "acquiringFeeVat": "1.5",
                    },
                ],
                "",
            )
        if payload.get("rrdId") == 101:
            return (
                200,
                [
                    {
                        "rrdId": 102,
                        "reportId": 11,
                        "documentType": "Возврат",
                        "acquiringFee": "3",
                        "acquiringFeeVat": "0.5",
                    }
                ],
                "",
            )
        return 204, None, ""

    result = build_month_acquiring_report(
        "token",
        2026,
        4,
        limiter=FinanceRateLimiter(0, sleep=lambda _s: None),
        post_fn=post_fn,
        page_size=2,
    )
    assert len(result.rows) == 3
    assert result.net_fee == Decimal("15")
    assert result.net_vat == Decimal("3.0")
    assert result.filename == "wb_acquiring_2026-04.xlsx"
    assert any("/acquiring/list" in url for url, _payload in calls)
    assert sum(1 for url, _payload in calls if "/acquiring/detailed/11" in url) == 2


def test_build_month_acquiring_report_empty() -> None:
    def post_fn(url: str, payload: dict):
        return 204, None, ""

    try:
        build_month_acquiring_report(
            "token",
            2026,
            4,
            limiter=FinanceRateLimiter(0, sleep=lambda _s: None),
            post_fn=post_fn,
        )
    except ValueError as exc:
        assert "нет" in str(exc).casefold()
    else:
        raise AssertionError("expected ValueError")
