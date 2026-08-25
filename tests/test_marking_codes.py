"""Разбор Data Matrix Честного знака и сопоставление GTIN с артикулом."""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from app.catalog_repository import CatalogRepository
from app.crm_repository import CrmRepository
from app.marking.cis import GS, parse_cis, split_cis_input
from app.marking.export import build_marking_codes_export
from app.marking.gtin import pad_gtin14
from app.marking.match import match_datamatrix_codes


GTIN13 = "4600605012345"
GTIN14 = "04600605012345"


def _repo(tmp_path, name: str = "marking.db") -> CatalogRepository:
    db_url = f"sqlite:///{(tmp_path / name).as_posix()}"
    crm = CrmRepository(db_url)
    crm.init_schema()
    repo = CatalogRepository(db_url)
    repo.init_schema()
    return repo


def _cis(gtin14: str, serial: str = "Ab12Xy", *, with_gs: bool = True) -> str:
    body = f"01{gtin14}21{serial}"
    crypto = "91FFD092abcdef0123456789"
    if with_gs:
        return f"{body}{GS}91FFD0{GS}92abcdef0123456789"
    return body + crypto


def test_pad_gtin13_to_gtin14() -> None:
    assert pad_gtin14(GTIN13) == GTIN14


def test_parse_cis_with_gs_extracts_gtin_and_serial() -> None:
    raw = _cis(GTIN14, with_gs=True)
    rec = parse_cis(raw)
    assert rec.ok
    assert rec.gtin == GTIN14
    assert rec.serial == "Ab12Xy"


def test_parse_cis_without_gs_uses_crypto_marker() -> None:
    rec = parse_cis(_cis(GTIN14, with_gs=False))
    assert rec.ok
    assert rec.gtin == GTIN14
    assert rec.serial == "Ab12Xy"


def test_parse_cis_parentheses_and_aim_and_placeholder() -> None:
    raw = f"]d2(01){GTIN14}(21)SER01(91)FFD0(92)aabb"
    rec = parse_cis(raw)
    assert rec.gtin == GTIN14
    assert rec.serial == "SER01"

    rec2 = parse_cis(f"01{GTIN14}21SER02<GS>91FFD0<GS>92aabb")
    assert rec2.gtin == GTIN14
    assert rec2.serial == "SER02"


def test_bare_gtin_is_not_datamatrix() -> None:
    rec = parse_cis(GTIN13)
    assert not rec.ok
    assert "GTIN" in rec.error


def test_split_skips_blank_and_comments() -> None:
    text = f"# header\n\n{_cis(GTIN14)}\n"
    lines = split_cis_input(text)
    assert lines == [_cis(GTIN14)]


def test_match_by_explicit_gtin_and_export_excel(tmp_path) -> None:
    repo = _repo(tmp_path)
    repo.create_product(
        {
            "name": "Болт",
            "sku": "BOLT-1",
            "code": "00001",
            "is_kit": False,
            "gtins": [GTIN13],
            "barcodes": [],
            "components": [],
        }
    )
    code_a = _cis(GTIN14, "AAA111")
    code_b = _cis(GTIN14, "BBB222")
    result = match_datamatrix_codes(f"{code_a}\n{code_b}\n{code_a}\n", repo)
    assert result.duplicate_count == 1
    assert len(result.groups) == 1
    group = result.groups[0]
    assert group.product.sku == "BOLT-1"
    assert group.gtin == GTIN14
    assert group.codes == [code_a, code_b]

    raw = build_marking_codes_export(result)
    wb = load_workbook(BytesIO(raw))
    grouped = wb["По товарам"]
    assert grouped.cell(1, 1).value == "Артикул"
    assert grouped.cell(2, 1).value == "BOLT-1"
    assert grouped.cell(2, 2).value == GTIN14
    assert "AAA111" in str(grouped.cell(2, 3).value)
    assert "BBB222" in str(grouped.cell(2, 3).value)
    codes_sheet = wb["Коды"]
    assert codes_sheet.cell(2, 1).value == "BOLT-1"
    assert codes_sheet.cell(2, 3).value
    assert codes_sheet.cell(3, 3).value


def test_match_by_ean13_barcode(tmp_path) -> None:
    repo = _repo(tmp_path, "ean.db")
    repo.create_product(
        {
            "name": "Гайка",
            "sku": "NUT-1",
            "code": "00002",
            "is_kit": False,
            "barcodes": [{"barcode": GTIN13, "label": "", "group": ""}],
            "components": [],
        }
    )
    result = match_datamatrix_codes(_cis(GTIN14, "ZX9"), repo)
    assert len(result.groups) == 1
    assert result.groups[0].product.sku == "NUT-1"
    assert result.unmatched == []


def test_unmatched_gtin_and_invalid(tmp_path) -> None:
    repo = _repo(tmp_path, "miss.db")
    repo.create_product(
        {
            "name": "Другой",
            "sku": "OTHER",
            "code": "00003",
            "is_kit": False,
            "gtins": ["4601111111111"],
            "components": [],
        }
    )
    text = f"{_cis(GTIN14)}\nnot-a-code\n"
    result = match_datamatrix_codes(text, repo)
    assert result.groups == []
    assert len(result.unmatched) == 1
    assert result.unmatched[0].gtin == GTIN14
    assert len(result.invalid) == 1


def test_gtin_unique_across_products(tmp_path) -> None:
    repo = _repo(tmp_path, "uniq.db")
    repo.create_product(
        {
            "name": "A",
            "sku": "A-1",
            "code": "00011",
            "is_kit": False,
            "gtins": [GTIN14],
            "components": [],
        }
    )
    try:
        repo.create_product(
            {
                "name": "B",
                "sku": "B-1",
                "code": "00012",
                "is_kit": False,
                "gtins": [GTIN13],
                "components": [],
            }
        )
    except ValueError as exc:
        assert "GTIN" in str(exc)
    else:
        raise AssertionError("ожидали ошибку уникальности GTIN")


def test_add_and_remove_gtin(tmp_path) -> None:
    repo = _repo(tmp_path, "addgtin.db")
    product = repo.create_product(
        {
            "name": "Болт",
            "sku": "BOLT-1",
            "code": "00001",
            "is_kit": False,
            "components": [],
        }
    )
    assert repo.add_product_gtin(int(product.id), GTIN13) == "created"
    assert repo.add_product_gtin(int(product.id), GTIN14) == "exists"
    rows = repo.list_product_gtin_rows()
    assert rows[0]["gtins"] == [GTIN14]
    assert repo.remove_product_gtin(int(product.id), GTIN14) is True
    assert repo.list_product_gtin_rows()[0]["gtins"] == []


def test_gtin_excel_import(tmp_path) -> None:
    from app.marking.gtin_import import build_gtin_import_template, import_gtins_from_xlsx

    repo = _repo(tmp_path, "imp.db")
    repo.create_product(
        {
            "name": "Болт",
            "sku": "BOLT-1",
            "code": "00001",
            "is_kit": False,
            "components": [],
        }
    )
    raw = build_gtin_import_template(repo)
    wb = load_workbook(BytesIO(raw))
    ws = wb.active
    assert ws.cell(1, 4).value == "GTIN*"
    ws.cell(2, 4).value = GTIN13
    buf = BytesIO()
    wb.save(buf)
    result = import_gtins_from_xlsx(repo, buf.getvalue())
    assert result.created == 1
    assert result.failed == 0
    assert repo.list_product_gtin_rows()[0]["gtins"] == [GTIN14]


def test_marking_http_gtin_and_scan_export(tmp_path) -> None:
    from fastapi import FastAPI
    from fastapi.testclient import TestClient

    from app.web.warehouse_marking_routes import register_warehouse_marking_routes
    from app.warehouse_users_repository import WarehouseUserRow

    repo = _repo(tmp_path, "http2.db")
    product = repo.create_product(
        {
            "name": "Болт",
            "sku": "BOLT-1",
            "code": "00001",
            "is_kit": False,
            "components": [],
        }
    )
    app = FastAPI()
    fake_user = WarehouseUserRow(
        id=1,
        login="admin",
        display_name="Admin",
        group_id=None,
        group_name="",
        telegram_nick="",
        is_admin=True,
        is_active=True,
        permissions={},
        created_at_ts=0,
        updated_at_ts=0,
    )

    def require_warehouse_user() -> WarehouseUserRow:
        return fake_user

    register_warehouse_marking_routes(app, repo, require_warehouse_user)
    client = TestClient(app)
    added = client.post(
        "/api/warehouse/marking/gtins",
        json={"product_id": product.id, "gtin": GTIN13},
    )
    assert added.status_code == 200, added.text
    listed = client.get("/api/warehouse/marking/gtins")
    assert listed.status_code == 200
    assert listed.json()["products"][0]["gtins"] == [GTIN14]

    exported = client.post(
        "/api/warehouse/marking/codes/export",
        json={"codes": [_cis(GTIN14, "AAA111"), _cis(GTIN14, "BBB222")]},
    )
    assert exported.status_code == 200, exported.text
    wb = load_workbook(BytesIO(exported.content))
    assert wb["Коды"].cell(2, 1).value == "BOLT-1"
    assert wb["Коды"].max_row == 3
